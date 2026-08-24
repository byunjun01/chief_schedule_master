import streamlit as st
import pandas as pd
import json
import io
import os
import re
import math
import traceback
from datetime import datetime, timedelta
from utils import PROF_ORDER, RAW_SCHEDULES_INITIAL, get_task_style, get_prof_raw_style, generate_schedule, run_auto_assignment, biweekly_week_active, is_prof_off
from manual import show_manual # [추가] 사용법 모듈 불러오기
from verification import verify_schedule  # [추가] 스케줄 검증 모듈
from cpsat_solver import DISABLEABLE_RULES  # 하드 룰 임시 해제 UI용 목록

st.set_page_config(page_title="의국 스케줄 마스터", layout="wide")

# --- 초기 상태 설정 ---
if 'base_date' not in st.session_state: st.session_state.base_date = datetime(2026, 6, 1).date()
if 'week_count' not in st.session_state: st.session_state.week_count = 5
# 단독 격주 진료의 생성 주차 선택 {"교수|요일|시간|진료명": 'odd'/'even'}. 매달 달력 따라 사용자가 지정.
if 'biweekly_choice' not in st.session_state: st.session_state.biweekly_choice = {}
# 사용자가 체크 해제한 하드 룰 key 목록 (기본: 전부 적용 = 빈 리스트)
if 'disabled_rules' not in st.session_state: st.session_state.disabled_rules = []
# 사용자가 직접 못박은 배정 [{'date','task','person'}, ...]
if 'manual_pins' not in st.session_state: st.session_state.manual_pins = []
if 'user_holidays_str' not in st.session_state: st.session_state.user_holidays_str = ""
if 'off_slots' not in st.session_state: st.session_state.off_slots = []
if 'residents' not in st.session_state: st.session_state.residents = []
if 'resident_leaves' not in st.session_state: st.session_state.resident_leaves = []
if 'pain_applicants' not in st.session_state: st.session_state.pain_applicants = []
if 'student_practices' not in st.session_state: st.session_state.student_practices = []
if 'assignments' not in st.session_state: st.session_state.assignments = {}
if 'cpsat_skipped_obs' not in st.session_state: st.session_state.cpsat_skipped_obs = []  # 제외룰로 드롭된 판정참관 task_id
if 'cpsat_skipped_prev_month_obs' not in st.session_state: st.session_state.cpsat_skipped_prev_month_obs = []  # 이전달 참관 중 미배정된 것
if 'relax_prev_month_obs' not in st.session_state: st.session_state.relax_prev_month_obs = False  # 이전달 참관 미배정 페널티 완화 여부
if 'drop_prev_month_obs' not in st.session_state: st.session_state.drop_prev_month_obs = False  # 이전달 참관 아예 배정 안 함
# 연차별 판정 개수 범위 {'R2':[lo,hi],'R1':[lo,hi],'R0':[lo,hi],'week_max':n}
if 'panjung_ranges' not in st.session_state: st.session_state.panjung_ranges = {}
# 장기휴가자(가용 절반 이하) 로딩 범위
if 'longleave_range' not in st.session_state: st.session_state.longleave_range = [0.0, 9.0]
# 보건소/영상 주당 추가 배정
if 'weekly_extra_bogeonso' not in st.session_state: st.session_state.weekly_extra_bogeonso = False
if 'weekly_extra_rad' not in st.session_state: st.session_state.weekly_extra_rad = False
if 'alloc_report' not in st.session_state: st.session_state.alloc_report = ""
if 'current_df_all' not in st.session_state: st.session_state.current_df_all = pd.DataFrame()
if 'res_daily_slots' not in st.session_state: st.session_state.res_daily_slots = {}
if 'bogeonso_substitutes' not in st.session_state: st.session_state.bogeonso_substitutes = {}  # {"MM-DD": [name1, name2, ...]}
if 'supplementary_schedules' not in st.session_state: st.session_state.supplementary_schedules = []  # [{"교수","날짜","시간","진료명"}, ...]
if 'use_cpsat' not in st.session_state: st.session_state.use_cpsat = True  # CP-SAT 모드 (신규, 기본값)
if 'cpsat_time_limit' not in st.session_state: st.session_state.cpsat_time_limit = 60
if 'cpsat_manual_multiplier' not in st.session_state: st.session_state.cpsat_manual_multiplier = "자동"
# 로딩 범위 기본값 (그룹 0~4의 (하한, 상한)) — 사용자가 설정해 둔 기본설정
if 'loading_ranges' not in st.session_state:
    st.session_state.loading_ranges = {0: [4.9, 5.5], 1: [6.3, 6.8], 2: [6.5, 7.2], 3: [7.3, 7.9], 4: [7.5, 8.0], 5: [8.0, 9.0]}
# H17 부등호 기본값 (경계 0~4: 그룹0<1, 1<2, 2<3, 3<4, 4<5) — 모두 strict '<'
if 'h17_ops' not in st.session_state:
    st.session_state.h17_ops = {0: "<", 1: "<", 2: "<", 3: "<", 4: "<"}
# 깨도 되는 pairing 최대 개수 (기본 5)
if 'cpsat_max_broken_pairs' not in st.session_state:
    st.session_state.cpsat_max_broken_pairs = 5
# 슬롯부족이 아니어도 -1 이동 허용할 차리/판정 추가 개수 (기본 0)
if 'cpsat_extra_shift' not in st.session_state:
    st.session_state.cpsat_extra_shift = 0
# 교수 중복 최소화 (S5) ε-constraint 옵션
if 'cpsat_prof_repeat_enabled' not in st.session_state:
    st.session_state.cpsat_prof_repeat_enabled = False
if 'cpsat_prof_repeat_multiplier' not in st.session_state:
    st.session_state.cpsat_prof_repeat_multiplier = 1.20
# 최소 여유 점수 (cap = max(⌈배수×L⌉, L+slack) — 작은 L에서 무의미해지는 거 보완)
if 'cpsat_prof_repeat_slack' not in st.session_state:
    st.session_state.cpsat_prof_repeat_slack = 5
# Shift 금지 날짜 (MM-DD 문자열 또는 None) — 이 날짜로 차리/판정이 -1 이동되는 것 방지
if 'shift_forbid_date' not in st.session_state:
    st.session_state.shift_forbid_date = None
if 'skip_panjung_obs' not in st.session_state:
    st.session_state.skip_panjung_obs = []
# 엑셀 다운로드 시 빈 5주차 padding 여부 (스케줄이 4주여도 5주짜리 엑셀 양식 유지)
if 'excel_pad_week5' not in st.session_state:
    st.session_state.excel_pad_week5 = False
# 빈 5주차 체크박스는 두 탭에 하나씩 있어 위젯 key가 분리됨 → 공용 값으로 초기화
for _k in ('excel_pad_week5_ck_1', 'excel_pad_week5_ck_2'):
    if _k not in st.session_state:
        st.session_state[_k] = bool(st.session_state.excel_pad_week5)

if 'master_schedules' not in st.session_state:
    st.session_state.master_schedules = pd.DataFrame(RAW_SCHEDULES_INITIAL, columns=["교수명", "요일", "시간", "진료명", "주기", "차리생성", "참관생성", "태그"])

def _pins_to_task_ids(df, pins):
    """사용자 지정 [{date,task,person}] → {task_id: person}.
    task_id는 스케줄 재생성 때마다 바뀌므로 (날짜+업무명)으로 매번 다시 찾는다."""
    if not pins or df is None or df.empty:
        return None
    out = {}
    for p in pins:
        m = df[(df['date'] == p.get('date')) & (df['task'] == p.get('task'))]
        for tid in m['task_id'].tolist():
            out[tid] = p.get('person')
    return out or None

def get_date_options(base_date, week_count, extended=False):
    """
    extended=False (기본): 현재 주차 범위까지의 평일만 반환
    extended=True: 다음 1주 평일도 추가 (라벨에 [다음주] 태그)
        - 휴진/보충진료처럼 표시 범위 밖이지만 차리/판정 생성에 영향을 주는 항목에서 사용
    """
    dates = []
    weekdays = ["월", "화", "수", "목", "금"]
    for w in range(week_count):
        for d in range(5):
            dt = base_date + timedelta(days=w*7 + d)
            dates.append(f"{dt.strftime('%m-%d')} ({weekdays[d]})")
    if extended:
        # 다음 1주 평일 추가
        for d in range(5):
            dt = base_date + timedelta(days=week_count*7 + d)
            dates.append(f"{dt.strftime('%m-%d')} ({weekdays[d]}) [다음주]")
    return dates

def get_sorted_residents(residents):
    def rank_score(r):
        year, roles = r['연차'], r['역할']
        yp = {"R3": 400, "R2": 300, "R1": 200, "R0": 100}.get(year, 0)
        rp = 0
        if year == "R3":
            if "의국수석" in roles: rp = 60
            elif "교육수석" in roles: rp = 50
            elif "학생수석" in roles: rp = 40
            elif "진료수석" in roles: rp = 30
            elif "본원 영상" in roles: rp = 10
            else: rp = 20
        elif year == "R2": rp = 10 if "연건 보건소" in roles else 20
        return yp + rp
    return sorted(residents, key=rank_score, reverse=True)

# --- 엑셀 생성 함수 ---
def generate_excel_data(week_count, base_date, sorted_res_list, user_holidays, res_daily_slots, assignments, df_all, task_map, include_blank_week5=False):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "전공의 스케줄"
    font_default = Font(name='맑은 고딕', size=11); font_white = Font(name='맑은 고딕', size=11, color="FFFFFFFF")
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    side_thin = Side(style='thin'); side_thick = Side(style='medium')
    thin_border = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    thick_top_border = Border(left=side_thin, right=side_thin, top=side_thick, bottom=side_thin)
    ws.column_dimensions['A'].width = 10; ws.column_dimensions['B'].width = 15
    for col in ['C', 'D', 'E', 'F', 'G']: ws.column_dimensions[col].width = 28

    # === R3/R2 슬롯 8명씩 보장 — 실제 인원 부족 시 빈 슬롯(None)으로 패딩 ===
    r3_list = [r for r in sorted_res_list if r['연차'] == "R3"]
    r2_list = [r for r in sorted_res_list if r['연차'] == "R2"]
    other_list = [r for r in sorted_res_list if r['연차'] not in ["R3", "R2"]]
    while len(r3_list) < 8: r3_list.append(None)
    while len(r2_list) < 8: r2_list.append(None)
    # (그룹라벨, res_or_None) 슬롯 시퀀스
    week_slots = []
    for r in r3_list: week_slots.append(("R3", r))
    for r in r2_list: week_slots.append(("R2", r))
    for r in other_list:
        grp = "R1/R0" if r['연차'] in ["R1", "R0"] else r['연차']
        week_slots.append((grp, r))

    current_row = 1
    render_weeks = max(week_count, 5) if include_blank_week5 else week_count
    for w in range(1, render_weeks + 1):
        ws.cell(row=current_row, column=2).border = thin_border
        for d_idx, day_name in enumerate(["월", "화", "수", "목", "금"]):
            dt = base_date + timedelta(days=(w-1)*7 + d_idx)
            cell = ws.cell(row=current_row, column=3+d_idx, value=dt.strftime("%m월 %d일"))
            cell.font = font_default; cell.alignment = align_center; cell.border = thin_border; cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        start_row_for_week = current_row + 1; current_row += 1
        prev_group = None
        for grp, res in week_slots:
            r1, r2 = current_row, current_row + 1
            is_new_group = prev_group is not None and grp != prev_group
            prev_group = grp
            current_r1_border = thick_top_border if is_new_group else thin_border

            if res is None:
                # 빈 슬롯 (R3/R2 패딩) — 이름/역할/요일 칸 모두 빈 칸 + 테두리만
                for col_n in range(2, 8):
                    c1 = ws.cell(row=r1, column=col_n, value="")
                    c1.border = current_r1_border
                    c2 = ws.cell(row=r2, column=col_n, value="")
                    c2.border = thin_border
                current_row += 2
                continue

            res_name = res['이름']; roles = ", ".join(res['역할']) if res['역할'] else ""
            cell_name = ws.cell(row=r1, column=2, value=res_name); cell_role = ws.cell(row=r2, column=2, value=roles)
            cell_name.font = font_default; cell_name.alignment = align_center; cell_name.border = current_r1_border
            cell_role.font = font_default; cell_role.alignment = align_center; cell_role.border = thin_border
            for d_idx, day_name in enumerate(["월", "화", "수", "목", "금"]):
                d_str = (base_date + timedelta(days=(w-1)*7 + d_idx)).strftime("%m-%d")
                am_task, pm_task, am_bg, pm_bg, am_fg, pm_fg = "", "", "FFFFFF", "FFFFFF", "black", "black"
                if d_str in user_holidays:
                    am_task = pm_task = "공휴일"; am_bg = pm_bg = "A6A6A6"
                else:
                    a_id = res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get('오전')
                    if a_id:
                        if any(kw in a_id for kw in ["휴가", "Off"]): am_task = a_id; am_bg = "A6A6A6"; am_fg = "white"
                        elif a_id == "영상": am_task = a_id; am_bg = "D5E8D4"
                        elif any(kw in a_id for kw in ["메인외래", "연보"]): am_task = a_id; am_bg = "E8F8F5" if "메인" in a_id else "FF85FF"
                        else: am_task = task_map.get(a_id, a_id); am_bg, am_fg = get_task_style(am_task)
                    p_id = res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get('오후')
                    if p_id:
                        if any(kw in p_id for kw in ["휴가", "Off"]): pm_task = p_id; pm_bg = "A6A6A6"; pm_fg = "white"
                        elif p_id == "영상": pm_task = p_id; pm_bg = "D5E8D4"
                        elif any(kw in p_id for kw in ["메인외래", "연보"]): pm_task = p_id; pm_bg = "E8F8F5" if "메인" in p_id else "FF85FF"
                        else: pm_task = task_map.get(p_id, p_id); pm_bg, pm_fg = get_task_style(pm_task)
                am_bg_h = am_bg.replace("#", ""); pm_bg_h = pm_bg.replace("#", "")
                cell_am = ws.cell(row=r1, column=3+d_idx, value=am_task); cell_pm = ws.cell(row=r2, column=3+d_idx, value=pm_task)
                cell_am.font = font_white if am_fg == "white" else font_default; cell_pm.font = font_white if pm_fg == "white" else font_default
                cell_am.alignment = align_center; cell_pm.alignment = align_center
                cell_am.border = current_r1_border; cell_pm.border = thin_border
                if am_bg_h != "FFFFFF": cell_am.fill = PatternFill(start_color=am_bg_h, end_color=am_bg_h, fill_type="solid")
                if pm_bg_h != "FFFFFF": cell_pm.fill = PatternFill(start_color=pm_bg_h, end_color=pm_bg_h, fill_type="solid")
            current_row += 2
        ws.merge_cells(start_row=start_row_for_week-1, start_column=1, end_row=current_row-1, end_column=1)
        cw = ws.cell(row=start_row_for_week-1, column=1, value=f"{w}주차"); cw.font = font_default; cw.alignment = align_center; cw.border = thin_border
    output = io.BytesIO(); wb.save(output); return output.getvalue()


def generate_prof_schedule_excel(week_count, base_date, master_schedules, supplementary_schedules, user_holidays, off_slots, prof_order, include_blank_week5=False, biweekly_choice=None):
    """교수별 시간표 엑셀 — example.xlsx 양식과 동일.

    양식 디테일 (example 기준):
      - 폰트: 본문/헤더 11pt 맑은 고딕, 제목 14pt Bold, 헤더 Bold + 왼쪽 정렬
      - 테두리: 페어 박스 외곽 medium, 내부 thin / 범례 medium 사방
      - 행 높이: 제목 20.65, 범례·빈 행 17.25
      - 셀 색: 외래 #70AD47 / 건증 #FFD966 / 판정·클리닉 #FFF2CC / 참관만 #D9EAD3 / 그 외 흰배경
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "교수별 시간표"

    FONT_NAME = '맑은 고딕'
    font_body = Font(name=FONT_NAME, size=11)
    font_body_bold = Font(name=FONT_NAME, size=11, bold=True)
    font_title = Font(name=FONT_NAME, size=14, bold=True)
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _argb(rgb6):
        h = (rgb6 or '').replace('#', '').upper()
        if len(h) == 6: return 'FF' + h
        if len(h) == 8: return h
        return 'FFFFFFFF'

    def _cat_color(prof, ci, has_chari, has_chamgwan):
        if not ci: return None
        # 차리/참관 둘 다 미생성 → task 0 → 흰배경 (재택외래 포함)
        if not has_chari and not has_chamgwan:
            return None
        # 진료명이 '클리닉' 자체인 경우만 클리닉 색
        if '클리닉' in ci or '통증' in ci:
            return _argb('FFF2CC')
        if '판정' in ci and '건증' not in ci:
            return _argb('FFF2CC')   # 단독 판정 (민경하 등)
        if '건증' in ci:
            return _argb('FFD966')
        if '암외래' in ci or '외래' in ci:
            if has_chari and has_chamgwan:
                return _argb('70AD47')
            return _argb('D9EAD3')   # 차리만 또는 참관만
        return None

    def _b(top='thin', bottom='thin', left='thin', right='thin'):
        """Border 헬퍼 — 각 변마다 'medium' 또는 'thin' 선택."""
        return Border(
            top=Side(style=top), bottom=Side(style=bottom),
            left=Side(style=left), right=Side(style=right),
        )

    def _get_cell_text_and_color(p_name, day, time, w, d_str_mmdd):
        is_h = d_str_mmdd in user_holidays
        is_o = is_prof_off(off_slots, p_name, d_str_mmdd, time)
        ci = ''
        has_chari = True; has_chamgwan = True
        if master_schedules is not None and not master_schedules.empty:
            for _, r in master_schedules.iterrows():
                if pd.isna(r.get('교수명')): continue
                if r['교수명'] == p_name and r['요일'] == day and r['시간'] == time:
                    period = r.get('주기', '매주')
                    _clv = str(r['진료명']) if pd.notna(r['진료명']) else ""
                    if biweekly_week_active(str(period), w, f"{p_name}|{day}|{time}|{_clv}", biweekly_choice):
                        ci = r['진료명']
                        has_chari = bool(r.get('차리생성', True))
                        has_chamgwan = bool(r.get('참관생성', True))
        if not ci:
            for s in supplementary_schedules:
                if s.get('교수') == p_name and s.get('날짜') == d_str_mmdd and s.get('시간') == time:
                    ci = s.get('진료명', '')
                    has_chari = True; has_chamgwan = True
                    break
        if is_h: return '공휴일', None
        if is_o: return '휴진', None
        if ci: return ci, _cat_color(p_name, ci, has_chari, has_chamgwan)
        return '', None

    days_kor = ['월', '화', '수', '목', '금']

    # 컬럼 폭 (example은 디폴트 유지지만 진료명 가독성 위해 약간 늘림)
    ws.column_dimensions['A'].width = 6
    for c in 'BCDEF': ws.column_dimensions[c].width = 12
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 6
    for c in 'IJKLM': ws.column_dimensions[c].width = 12

    # === 최상단 제목 (행 3) ===
    start_d = base_date
    end_d = base_date + timedelta(days=week_count * 7 - 1)
    title = f"전공의 배정 현황 ({start_d.month}월, {start_d.strftime('%m/%d')} ~ {end_d.strftime('%m/%d')})"
    ws.merge_cells('A3:M3')
    c_t = ws.cell(row=3, column=1, value=title)
    c_t.font = font_title; c_t.alignment = align_c
    ws.row_dimensions[3].height = 20.65

    # === 범례 (행 5) — 사방 medium 테두리 ===
    legend = [
        (2, 3, '외래 차리/참관', '70AD47'),
        (4, 5, '건증 참관+판정', 'FFD966'),
        (6, 7, '외래 참관만', 'D9EAD3'),
        (8, 9, '건증 판정만', 'FFF2CC'),
        (10, 11, '전공의 배정 제외', None),
    ]
    for col1, col2, label, color in legend:
        ws.merge_cells(start_row=5, start_column=col1, end_row=5, end_column=col2)
        # 병합 영역의 각 셀에 medium 테두리
        for cc in range(col1, col2 + 1):
            cell = ws.cell(row=5, column=cc)
            top = bottom = 'medium'
            left = 'medium' if cc == col1 else None
            right = 'medium' if cc == col2 else None
            cell.border = Border(
                top=Side(style=top), bottom=Side(style=bottom),
                left=Side(style=left) if left else Side(),
                right=Side(style=right) if right else Side(),
            )
        c = ws.cell(row=5, column=col1, value=label)
        c.font = font_body; c.alignment = align_c
        if color:
            argb = _argb(color)
            c.fill = PatternFill(start_color=argb, end_color=argb, fill_type='solid')
    ws.row_dimensions[5].height = 17.25

    # === 교수 페어 블록 ===
    prof_list = list(prof_order)
    current_row = 8

    def _box_border(r, c, box_top, box_bottom, box_left, box_right):
        """페어 박스 안에서 (r, c)가 외곽인지 여부 → 해당 변 medium, 나머지 thin"""
        return _b(
            top='medium' if r == box_top else 'thin',
            bottom='medium' if r == box_bottom else 'thin',
            left='medium' if c == box_left else 'thin',
            right='medium' if c == box_right else 'thin',
        )

    render_weeks = max(week_count, 5) if include_blank_week5 else week_count
    for pair_idx in range(0, len(prof_list), 2):
        left_prof = prof_list[pair_idx]
        right_prof = prof_list[pair_idx + 1] if pair_idx + 1 < len(prof_list) else None

        box_top = current_row                          # 헤더 행
        box_bottom = current_row + render_weeks * 3    # 마지막 본문 행 (헤더1 + 주차×3, 0-index니까 +render_weeks*3)

        # === 헤더 행 (current_row) ===
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cL = ws.cell(row=current_row, column=1, value=f"{left_prof}교수님")
        cL.font = font_body_bold; cL.alignment = align_left
        # 왼쪽 박스 헤더 행 각 셀의 외곽 medium
        for cc in range(1, 7):
            ws.cell(row=current_row, column=cc).border = _box_border(current_row, cc, box_top, box_bottom, 1, 6)
        if right_prof:
            ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=13)
            cR = ws.cell(row=current_row, column=8, value=f"{right_prof}교수님")
            cR.font = font_body_bold; cR.alignment = align_left
            for cc in range(8, 14):
                ws.cell(row=current_row, column=cc).border = _box_border(current_row, cc, box_top, box_bottom, 8, 13)
        current_row += 1

        # === 주차 × (날짜 + 오전 + 오후) ===
        for w in range(1, render_weeks + 1):
            # 날짜 행
            for d_idx in range(5):
                dt = base_date + timedelta(days=(w - 1) * 7 + d_idx)
                d_label = dt.strftime('%m/%d')
                # 왼쪽 B-F
                c = ws.cell(row=current_row, column=2 + d_idx, value=d_label)
                c.font = font_body; c.alignment = align_c
                c.border = _box_border(current_row, 2 + d_idx, box_top, box_bottom, 1, 6)
                # 오른쪽 I-M
                if right_prof:
                    c = ws.cell(row=current_row, column=9 + d_idx, value=d_label)
                    c.font = font_body; c.alignment = align_c
                    c.border = _box_border(current_row, 9 + d_idx, box_top, box_bottom, 8, 13)
            # A/H 빈 라벨 칸도 외곽 테두리 (left=medium)
            ws.cell(row=current_row, column=1).border = _box_border(current_row, 1, box_top, box_bottom, 1, 6)
            if right_prof:
                ws.cell(row=current_row, column=8).border = _box_border(current_row, 8, box_top, box_bottom, 8, 13)
            current_row += 1

            # 오전/오후
            for time in ['오전', '오후']:
                cAL = ws.cell(row=current_row, column=1, value=time)
                cAL.font = font_body; cAL.alignment = align_c
                cAL.border = _box_border(current_row, 1, box_top, box_bottom, 1, 6)
                if right_prof:
                    cAR = ws.cell(row=current_row, column=8, value=time)
                    cAR.font = font_body; cAR.alignment = align_c
                    cAR.border = _box_border(current_row, 8, box_top, box_bottom, 8, 13)
                for d_idx, day in enumerate(days_kor):
                    d_str_mmdd = (base_date + timedelta(days=(w - 1) * 7 + d_idx)).strftime('%m-%d')
                    # 왼쪽
                    val, color = _get_cell_text_and_color(left_prof, day, time, w, d_str_mmdd)
                    c = ws.cell(row=current_row, column=2 + d_idx, value=val)
                    c.font = font_body; c.alignment = align_c
                    c.border = _box_border(current_row, 2 + d_idx, box_top, box_bottom, 1, 6)
                    if color: c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    # 오른쪽
                    if right_prof:
                        val, color = _get_cell_text_and_color(right_prof, day, time, w, d_str_mmdd)
                        c = ws.cell(row=current_row, column=9 + d_idx, value=val)
                        c.font = font_body; c.alignment = align_c
                        c.border = _box_border(current_row, 9 + d_idx, box_top, box_bottom, 8, 13)
                        if color: c.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                current_row += 1

        # 페어 사이 빈 행
        ws.row_dimensions[current_row].height = 17.25
        current_row += 1

    output = io.BytesIO(); wb.save(output); return output.getvalue()


# --- 엑셀 지연 생성 헬퍼 ---
def _sync_pad5(src_key):
    """'빈 5주차' 체크박스는 두 탭에 하나씩 있어 key가 다름 → 한쪽을 바꾸면 반대쪽도 맞춘다."""
    val = bool(st.session_state.get(src_key, False))
    st.session_state.excel_pad_week5 = val
    for k in ('excel_pad_week5_ck_1', 'excel_pad_week5_ck_2'):
        if k != src_key:
            st.session_state[k] = val


def _excel_sig(*parts):
    """엑셀 입력값 지문. 달라지면 캐시된 엑셀을 버리고 다시 만들게 한다."""
    try:
        return hash(json.dumps(parts, sort_keys=True, default=str))
    except Exception:
        return None


def _excel_download_ui(slot, sig_key, data_key, sig, build_fn, dl_label, file_name, btn_key):
    """엑셀을 매 rerun마다 만들지 않고, [만들기]를 누른 뒤에만 생성/캐시해서 다운로드 제공.
    입력값(sig)이 바뀌면 캐시를 버리고 다시 [만들기]를 노출한다."""
    _MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    with slot:
        if st.session_state.get(sig_key) == sig and st.session_state.get(data_key):
            st.download_button(label=dl_label, data=st.session_state[data_key],
                               file_name=file_name, mime=_MIME, type="primary")
        else:
            if st.session_state.get(data_key):
                st.caption("⚠️ 내용이 바뀌었습니다 — 다시 만들어주세요.")
            if st.button("📊 엑셀 만들기", key=btn_key, type="primary"):
                with st.spinner("엑셀 생성 중..."):
                    st.session_state[data_key] = build_fn()
                    st.session_state[sig_key] = sig
                st.rerun()


# --- 사이드바 ---
with st.sidebar:
    st.header("⚙️ 기본 설정")
    new_base_date = st.date_input("시작 월요일", value=st.session_state.base_date)
    new_week_count = st.radio("주차 설정", [4, 5], index=0 if st.session_state.week_count == 4 else 1)

    # === 격주(단독) 진료 주차 설정 팝업 ===
    # 짝 없는 단독 격주 진료(예: 전혜령)는 이번 달 달력에 따라 몇 주차에 들어갈지 매번 달라짐 → 여기서 지정.
    # (조우현/김계형처럼 둘이 나눠 맡는 슬롯은 자동으로 매주 합쳐 생성되므로 여기 안 나옴)
    _bw_groups = {}
    _msched = st.session_state.master_schedules
    if _msched is not None and not _msched.empty:
        for _, _r in _msched.iterrows():
            if pd.isna(_r.get('교수명')) or not str(_r.get('교수명')).strip(): continue
            _cyc = str(_r.get('주기', '매주'))
            if _cyc in ('홀수주', '짝수주'):
                _k = (str(_r['요일']), str(_r['시간']), str(_r['진료명']) if pd.notna(_r['진료명']) else '')
                _bw_groups.setdefault(_k, {})[_cyc] = str(_r['교수명'])
    _solo_bw = []
    for (_dy, _tm, _cl), _v in _bw_groups.items():
        if not ('홀수주' in _v and '짝수주' in _v):
            for _cyc, _p in _v.items():
                _solo_bw.append((_p, _dy, _tm, _cl, _cyc))
    _solo_bw.sort()
    # 둘이 나눠 맡는 공유 격주 슬롯 (조우현/김계형): 홀라벨/짝라벨 교수 쌍
    _shared_bw = []
    for (_dy, _tm, _cl), _v in _bw_groups.items():
        if '홀수주' in _v and '짝수주' in _v:
            _shared_bw.append((_v['홀수주'], _v['짝수주'], _dy, _tm, _cl))
    _shared_bw.sort()
    # ※ 여기서 '현재 슬롯에 없는 키'를 지우지 않는다.
    #    규칙설정 탭에서 진료명/요일/시간을 잠깐 바꾸면 key가 달라지는데,
    #    그때 정리해버리면 저장해둔 주차 선택이 조용히 사라진다.
    #    (안 쓰는 키가 남아도 generate_schedule은 해당 key만 조회하므로 무해)
    if _solo_bw or _shared_bw:
        _bw_changed = False
        with st.popover("🔁 격주 진료 주차 설정", use_container_width=True):
            st.caption("격주 진료가 **이번 달 몇 주차**에 들어갈지 지정하세요. 달력마다 달라 생성 전 확인이 필요합니다.")
            for (_p, _dy, _tm, _cl, _cyc) in _solo_bw:
                _bkey = f"{_p}|{_dy}|{_tm}|{_cl}"
                _bdef = st.session_state.biweekly_choice.get(_bkey) or ('odd' if _cyc == '홀수주' else 'even')
                _bsel = st.radio(
                    f"{_p} · {_dy} {_tm} {_cl} (격주)",
                    options=['odd', 'even'],
                    index=0 if _bdef == 'odd' else 1,
                    format_func=lambda v: '1·3·5주차' if v == 'odd' else '2·4주차',
                    key=f"bwradio_{_bkey}",
                )
                if st.session_state.biweekly_choice.get(_bkey) != _bsel:
                    _bw_changed = True
                st.session_state.biweekly_choice[_bkey] = _bsel
            # 둘이 나눠 맡는 진료: 누가 1·3·5주인지 선택 (한쪽이 135면 다른쪽은 자동 24)
            for (_op, _ep, _dy, _tm, _cl) in _shared_bw:
                _bkey = f"{_op}|{_dy}|{_tm}|{_cl}"
                _bdef = st.session_state.biweekly_choice.get(_bkey) or 'odd'
                _bsel = st.radio(
                    f"{_op}/{_ep} · {_dy} {_tm} {_cl} (둘이 격주)",
                    options=['odd', 'even'],
                    index=0 if _bdef == 'odd' else 1,
                    format_func=lambda v, _o=_op, _e=_ep: (
                        f"{_o} 1·3·5주 / {_e} 2·4주" if v == 'odd' else f"{_e} 1·3·5주 / {_o} 2·4주"),
                    key=f"bwradio_{_bkey}",
                    help="합쳐진 task 이름은 그대로 유지되고, 이 선택은 '그 주 담당 교수의 휴진'을 정확히 반영하는 데 쓰입니다.",
                )
                if st.session_state.biweekly_choice.get(_bkey) != _bsel:
                    _bw_changed = True
                st.session_state.biweekly_choice[_bkey] = _bsel
        # 선택이 바뀌면 즉시 스케줄 재생성 (따로 '적용'을 누르지 않아도 반영되도록)
        # ※ st.rerun()은 쓰지 않는다 — 사이드바는 본문보다 먼저 실행되므로 아래에서
        #   current_df_all을 읽을 때 이번 렌더에 바로 반영된다. (rerun 시 무한 루프 위험)
        if _bw_changed and not st.session_state.current_df_all.empty:
            _uh = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
            st.session_state.current_df_all = generate_schedule(
                st.session_state.base_date, st.session_state.week_count, _uh,
                st.session_state.master_schedules, st.session_state.off_slots,
                supplementary_schedules=st.session_state.supplementary_schedules,
                biweekly_choice=st.session_state.biweekly_choice,
            )
        # 현재 적용 중인 격주 설정을 항상 표시 (백업 불러오기 등으로 바뀌어도 바로 보이게)
        _bw_now = []
        for (_p, _dy, _tm, _cl, _cyc) in _solo_bw:
            _v = st.session_state.biweekly_choice.get(f"{_p}|{_dy}|{_tm}|{_cl}") or ('odd' if _cyc == '홀수주' else 'even')
            _bw_now.append(f"{_p} {'1·3·5주' if _v == 'odd' else '2·4주'}")
        for (_op, _ep, _dy, _tm, _cl) in _shared_bw:
            _v = st.session_state.biweekly_choice.get(f"{_op}|{_dy}|{_tm}|{_cl}") or 'odd'
            _bw_now.append(f"{_op} {'1·3·5주' if _v == 'odd' else '2·4주'}/{_ep} {'2·4주' if _v == 'odd' else '1·3·5주'}")
        if _bw_now:
            st.caption("🔁 격주 현재 설정: " + " · ".join(_bw_now))

    st.markdown("---")
    st.header("🏖️ 교수 휴진/공휴일 관리")
    new_holidays_str = st.text_input("공휴일 (MM-DD, 쉼표 구분)", value=st.session_state.user_holidays_str)
    if st.button("📅 설정 및 공휴일 확정 적용", use_container_width=True, type="primary"):
        st.session_state.base_date = new_base_date
        st.session_state.week_count = new_week_count
        st.session_state.user_holidays_str = new_holidays_str
        u_holidays = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
        st.session_state.current_df_all = generate_schedule(new_base_date, new_week_count, u_holidays, st.session_state.master_schedules, st.session_state.off_slots, supplementary_schedules=st.session_state.supplementary_schedules, biweekly_choice=st.session_state.get('biweekly_choice'))
        st.success("설정이 스케줄에 반영되었습니다."); st.rerun()
    user_holidays = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
    available_dates = get_date_options(st.session_state.base_date, st.session_state.week_count)
    # 휴진/보충진료용: 다음 1주까지 포함 (표시 범위 밖이지만 차리/판정 생성에 영향)
    available_dates_extended = get_date_options(st.session_state.base_date, st.session_state.week_count, extended=True)
    with st.form("off_form", clear_on_submit=True):
        p_select = st.selectbox("교수님", PROF_ORDER)
        d_selects = st.multiselect("휴진 날짜", options=available_dates_extended, help="[다음주] 태그가 붙은 날짜는 표시 범위 밖이지만 차리/판정 생성에 영향을 줍니다")
        t_select = st.radio(
            "휴진 범위", ["종일", "오전", "오후"], horizontal=True,
            help="기본은 '종일'입니다. '오전'/'오후'를 고르면 그 시간대 진료만 휴진 처리되고, 반대 시간대 진료는 그대로 생성됩니다.",
        )
        if st.form_submit_button("휴진 등록"):
            for d_str in d_selects:
                # 라벨 형식: "MM-DD (요일)" 또는 "MM-DD (요일) [다음주]" → MM-DD만 추출
                date_only = d_str.split(" ")[0]
                _entry = (p_select, date_only, t_select)
                # 같은 교수·날짜의 기존 등록 정리: 종일을 새로 넣으면 반일 등록은 흡수
                _dups = [s for s in st.session_state.off_slots
                         if len(s) >= 2 and s[0] == p_select and s[1] == date_only]
                if t_select == "종일":
                    for _dup in _dups:
                        st.session_state.off_slots.remove(_dup)
                    st.session_state.off_slots.append(_entry)
                else:
                    # 이미 종일 휴진이면 반일 추가는 의미 없음
                    _has_full = any(len(s) < 3 or s[2] in ('', None, '종일') for s in _dups)
                    _same = any(len(s) >= 3 and s[2] == t_select for s in _dups)
                    if not _has_full and not _same:
                        st.session_state.off_slots.append(_entry)
            st.session_state.current_df_all = generate_schedule(st.session_state.base_date, st.session_state.week_count, user_holidays, st.session_state.master_schedules, st.session_state.off_slots, supplementary_schedules=st.session_state.supplementary_schedules, biweekly_choice=st.session_state.get('biweekly_choice')); st.rerun()
    if st.session_state.off_slots:
        with st.expander(f"📋 등록된 휴진 ({len(st.session_state.off_slots)}건)", expanded=True):
            # 교수별 → 날짜순 정렬하여 가독성 향상 (원래 인덱스는 삭제용으로 보존)
            indexed = sorted(
                enumerate(st.session_state.off_slots),
                key=lambda x: (PROF_ORDER.index(x[1][0]) if x[1][0] in PROF_ORDER else 999, x[1][1])
            )
            n_cols = 2
            grid_cols = st.columns(n_cols)
            for pos, (i, _slot) in enumerate(indexed):
                p, d = _slot[0], _slot[1]
                _tm = _slot[2] if len(_slot) >= 3 else '종일'
                if _tm in ('', None):
                    _tm = '종일'
                with grid_cols[pos % n_cols]:
                    cc1, cc2 = st.columns([4, 1])
                    _badge = '' if _tm == '종일' else f" · **{_tm}만**"
                    cc1.write(f"· {p} ({d}){_badge}")
                    if cc2.button("X", key=f"prof_off_{i}"):
                        st.session_state.off_slots.pop(i)
                        st.session_state.current_df_all = generate_schedule(st.session_state.base_date, st.session_state.week_count, user_holidays, st.session_state.master_schedules, st.session_state.off_slots, supplementary_schedules=st.session_state.supplementary_schedules, biweekly_choice=st.session_state.get('biweekly_choice'))
                        st.rerun()
    st.markdown("---")
    st.header("➕ 보충진료 추가")
    st.caption("단발성 진료를 추가하면 차리/판정/참관이 자동 생성됩니다. (위치: 진료일 -2영업일)")
    PROFS_WITH_AMOEHRAE = ["황서은", "조수환", "민경하", "김지영", "김하진"]
    PROFS_WITH_CLINIC = ["조비룡", "박민선", "박진호"]
    with st.form("supp_form", clear_on_submit=True):
        sup_prof = st.selectbox("교수님", PROF_ORDER, key="sup_prof")
        sup_dates = st.multiselect("날짜 (복수 선택 가능)", options=available_dates_extended, key="sup_dates", help="[다음주] 태그가 붙은 날짜는 표시 범위 밖이지만 차리/판정 생성에 영향을 줍니다")
        sup_time = st.selectbox("시간", ["오전", "오후"], key="sup_time")
        # 진료종류 동적 옵션
        clinic_options = ["건증", "외래"]
        if sup_prof in PROFS_WITH_AMOEHRAE:
            clinic_options.append("암외래")
        if sup_prof in PROFS_WITH_CLINIC:
            clinic_options.append("클리닉")
        sup_clinic = st.selectbox("진료종류", clinic_options, key="sup_clinic")
        if st.form_submit_button("➕ 보충진료 등록"):
            # 박진호 클리닉 → 통증클리닉으로 저장
            clinic_to_save = "통증클리닉" if (sup_prof == "박진호" and sup_clinic == "클리닉") else sup_clinic
            error_dates = []
            added_count = 0
            for d_str_full in sup_dates:
                d_str = d_str_full.split(" ")[0]
                # 휴진 충돌 검사
                if is_prof_off(st.session_state.off_slots, sup_prof, d_str, sup_time):
                    error_dates.append(d_str)
                    continue
                # 중복 검사
                already_exists = any(
                    s["교수"] == sup_prof and s["날짜"] == d_str and s["시간"] == sup_time and s["진료명"] == clinic_to_save
                    for s in st.session_state.supplementary_schedules
                )
                if already_exists:
                    continue
                st.session_state.supplementary_schedules.append({
                    "교수": sup_prof, "날짜": d_str, "시간": sup_time, "진료명": clinic_to_save
                })
                added_count += 1
            if error_dates:
                st.error(f"❌ {sup_prof} 교수는 다음 날짜에 휴진 등록되어 있어 보충진료를 추가할 수 없습니다: {', '.join(error_dates)}")
            if added_count > 0:
                st.session_state.current_df_all = generate_schedule(
                    st.session_state.base_date, st.session_state.week_count, user_holidays,
                    st.session_state.master_schedules, st.session_state.off_slots,
                    supplementary_schedules=st.session_state.supplementary_schedules,
                    biweekly_choice=st.session_state.get('biweekly_choice')
                )
                st.rerun()
    if st.session_state.supplementary_schedules:
        st.caption(f"등록된 보충진료 ({len(st.session_state.supplementary_schedules)}건):")
        for i, s in enumerate(st.session_state.supplementary_schedules):
            c1, c2 = st.columns([4, 1])
            c1.write(f"· {s['교수']} ({s['날짜']} {s['시간']} {s['진료명']})")
            if c2.button("X", key=f"sup_del_{i}"):
                st.session_state.supplementary_schedules.pop(i)
                st.session_state.current_df_all = generate_schedule(
                    st.session_state.base_date, st.session_state.week_count, user_holidays,
                    st.session_state.master_schedules, st.session_state.off_slots,
                    supplementary_schedules=st.session_state.supplementary_schedules,
                    biweekly_choice=st.session_state.get('biweekly_choice')
                )
                st.rerun()
    st.markdown("---")
    st.header("💾 백업 및 복구")
    export_data = {
        "base_date": new_base_date.strftime("%Y-%m-%d"),
        "week_count": new_week_count,
        "user_holidays_str": new_holidays_str,
        "off_slots": st.session_state.off_slots,
        "residents": st.session_state.residents,
        "resident_leaves": st.session_state.resident_leaves,
        "pain_applicants": st.session_state.pain_applicants,
        "student_practices": st.session_state.student_practices,
        "assignments": st.session_state.assignments,
        "bogeonso_substitutes": st.session_state.bogeonso_substitutes,
        "supplementary_schedules": st.session_state.supplementary_schedules,
        "master_schedules": st.session_state.master_schedules.to_dict(orient="records"),
        # 디버깅용 - CP-SAT 결과의 시간/날짜 결정 정보
        "task_time_overrides": st.session_state.get("cpsat_task_time_overrides", {}),
        "task_date_overrides": st.session_state.get("cpsat_task_date_overrides", {}),
        "shifted_tasks": st.session_state.get("cpsat_shifted_tasks", []),
        "original_df_dates": st.session_state.get("cpsat_original_df_dates", {}),
        # 배정 후 요약 (배정 리포트, S5 교수 중복 메타데이터)
        "alloc_report": st.session_state.get("alloc_report", ""),
        "prof_repeat_summary": st.session_state.get("prof_repeat_summary"),
        "prof_repeat_pairs_data": st.session_state.get("prof_repeat_pairs_data"),
        "prof_repeat_by_resident_data": st.session_state.get("prof_repeat_by_resident_data"),
        "shift_forbid_date": st.session_state.get("shift_forbid_date"),
        "excel_pad_week5": st.session_state.get("excel_pad_week5", False),
        "skip_panjung_obs": st.session_state.get("skip_panjung_obs", []),
        "cpsat_skipped_obs": st.session_state.get("cpsat_skipped_obs", []),
        "cpsat_skipped_prev_month_obs": st.session_state.get("cpsat_skipped_prev_month_obs", []),
        "relax_prev_month_obs": st.session_state.get("relax_prev_month_obs", False),
        "drop_prev_month_obs": st.session_state.get("drop_prev_month_obs", False),
        "panjung_ranges": st.session_state.get("panjung_ranges", {}),
        "longleave_range": st.session_state.get("longleave_range", [0.0, 9.0]),
        "weekly_extra_bogeonso": st.session_state.get("weekly_extra_bogeonso", False),
        "weekly_extra_rad": st.session_state.get("weekly_extra_rad", False),
        "biweekly_choice": st.session_state.get("biweekly_choice", {}),
        "disabled_rules": st.session_state.get("disabled_rules", []),
        "manual_pins": st.session_state.get("manual_pins", []),
        # === 스케줄 배정 탭의 솔버 설정 (이전엔 저장 안 되던 값들) ===
        "cpsat_time_limit": st.session_state.get("cpsat_time_limit", 60),
        "cpsat_manual_multiplier": st.session_state.get("cpsat_manual_multiplier", "자동"),
        "cpsat_max_broken_pairs": st.session_state.get("cpsat_max_broken_pairs", 5),
        "cpsat_extra_shift": st.session_state.get("cpsat_extra_shift", 0),
        "cpsat_prof_repeat_enabled": st.session_state.get("cpsat_prof_repeat_enabled", False),
        "cpsat_prof_repeat_multiplier": st.session_state.get("cpsat_prof_repeat_multiplier", 1.2),
        "cpsat_prof_repeat_slack": st.session_state.get("cpsat_prof_repeat_slack", 5),
        # 로딩 범위/부등호 (키가 int라 JSON에선 문자열로 저장 → 복구 시 int 변환)
        "loading_ranges": {str(g): list(v) for g, v in (st.session_state.get("loading_ranges") or {}).items()},
        "h17_ops": {str(g): v for g, v in (st.session_state.get("h17_ops") or {}).items()},
    }
    _export_json = json.dumps(export_data, ensure_ascii=False, indent=2)
    _default_name = f"backup_{datetime.today().strftime('%Y%m%d')}.json"

    # (1) 브라우저 다운로드 방식 (기존) — 저장 위치는 브라우저가 결정
    st.download_button("📥 설정 저장 (브라우저 다운로드)", _export_json, _default_name, "application/json", use_container_width=True)

    # (2) 이 PC의 지정 폴더에 직접 저장 — 시크릿/캐시와 무관하게 항상 같은 위치
    #     마지막 저장 폴더는 사이드카 파일(.last_backup_dir)에 기록해 재시작해도 기억
    _dir_memo = os.path.join(os.getcwd(), ".last_backup_dir")
    if 'backup_save_dir' not in st.session_state:
        _remembered = None
        try:
            if os.path.exists(_dir_memo):
                with open(_dir_memo, 'r', encoding='utf-8') as _f:
                    _remembered = _f.read().strip()
        except Exception:
            _remembered = None
        st.session_state.backup_save_dir = _remembered or os.path.join(os.getcwd(), "backups")

    with st.expander("💾 이 PC 폴더에 바로 저장 (시크릿 창에서도 항상 같은 위치)"):
        _save_dir = st.text_input("저장 폴더", value=st.session_state.backup_save_dir, key="backup_save_dir_input")
        _save_name = st.text_input("파일 이름", value=_default_name, key="backup_save_name_input")
        _full_path = os.path.join(_save_dir, _save_name) if _save_dir and _save_name else ""
        if _full_path and os.path.exists(_full_path):
            st.caption(f"⚠️ 이미 존재 — 저장 시 덮어씁니다: `{_full_path}`")
        if st.button("💾 이 폴더에 바로 저장", type="primary", use_container_width=True):
            try:
                os.makedirs(_save_dir, exist_ok=True)
                with open(_full_path, 'w', encoding='utf-8') as _f:
                    _f.write(_export_json)
                st.session_state.backup_save_dir = _save_dir
                try:
                    with open(_dir_memo, 'w', encoding='utf-8') as _f:
                        _f.write(_save_dir)   # 다음 실행에서도 이 폴더를 기본값으로
                except Exception:
                    pass
                st.success(f"✅ 저장 완료: {_full_path}")
            except Exception as _e:
                st.error(f"❌ 저장 실패: {_e}")
    uploaded_file = st.file_uploader("📤 불러오기 (파일 선택)", type=["json"])

    # 이 PC 폴더에서 직접 불러오기 — 마지막 불러온 폴더를 기억(.last_load_dir)해 처음에 그 폴더를 띄움
    _load_memo = os.path.join(os.getcwd(), ".last_load_dir")
    if 'backup_load_dir' not in st.session_state:
        _rem = None
        try:
            if os.path.exists(_load_memo):
                with open(_load_memo, 'r', encoding='utf-8') as _f:
                    _rem = _f.read().strip()
        except Exception:
            _rem = None
        st.session_state.backup_load_dir = _rem or st.session_state.get('backup_save_dir') or os.path.join(os.getcwd(), "backups")

    _disk_pick = None
    _disk_load_clicked = False
    _load_dir = st.session_state.backup_load_dir
    with st.expander("📂 이 PC 폴더에서 불러오기 (마지막 폴더가 처음에 뜸)"):
        _load_dir = st.text_input("불러올 폴더", value=st.session_state.backup_load_dir, key="backup_load_dir_input")
        _json_files = []
        if _load_dir and os.path.isdir(_load_dir):
            try:
                _json_files = sorted([f for f in os.listdir(_load_dir) if f.lower().endswith('.json')], reverse=True)
            except Exception as _e:
                st.error(f"폴더를 읽을 수 없습니다: {_e}")
        elif _load_dir:
            st.caption("⚠️ 폴더 경로가 올바르지 않습니다.")
        if _json_files:
            _disk_pick = st.selectbox("파일 선택 (최신순)", _json_files, key="backup_load_file_pick")
            _disk_load_clicked = st.button("📂 이 파일 불러오기", type="primary", use_container_width=True)
        elif _load_dir and os.path.isdir(_load_dir):
            st.caption("이 폴더에 .json 백업이 없습니다.")

    data = None
    if uploaded_file is not None and st.button("데이터 적용"):
        data = json.load(uploaded_file)
    elif _disk_pick and _disk_load_clicked:
        try:
            with open(os.path.join(_load_dir, _disk_pick), 'r', encoding='utf-8') as _f:
                data = json.load(_f)
            st.session_state.backup_load_dir = _load_dir
            try:
                with open(_load_memo, 'w', encoding='utf-8') as _f:
                    _f.write(_load_dir)   # 다음 실행에서도 이 폴더를 기본값으로
            except Exception:
                pass
        except Exception as _e:
            st.error(f"❌ 불러오기 실패: {_e}")
            data = None

    if data is not None:
        st.session_state.base_date = datetime.strptime(data["base_date"], "%Y-%m-%d").date()
        st.session_state.week_count = data.get("week_count", 5)
        st.session_state.user_holidays_str = data.get("user_holidays_str", "")
        st.session_state.off_slots = [tuple(x) for x in data.get("off_slots", [])]
        st.session_state.residents = data.get("residents", [])
        st.session_state.resident_leaves = data.get("resident_leaves", [])
        st.session_state.pain_applicants = data.get("pain_applicants", [])
        st.session_state.student_practices = data.get("student_practices", [])
        st.session_state.assignments = data.get("assignments", {})
        st.session_state.bogeonso_substitutes = data.get("bogeonso_substitutes", {})
        st.session_state.supplementary_schedules = data.get("supplementary_schedules", [])
        st.session_state.master_schedules = pd.DataFrame(data.get("master_schedules", RAW_SCHEDULES_INITIAL))
        # 격주 주차 선택은 스케줄 생성 '이전'에 복구해야 df에 바로 반영됨
        st.session_state.biweekly_choice = dict(data.get("biweekly_choice", {}) or {})
        for _wk in [k for k in list(st.session_state.keys()) if k.startswith('bwradio_')]:
            st.session_state.pop(_wk, None)
        # 전공의 명단 위젯 상태 정리 — 이름 기반 key가 이전 명단 값으로 남아 있으면
        # (같은 이름이 다른 연차로 바뀐 경우 등) 역할 multiselect가 options 불일치로 죽는다
        for _wk in [k for k in list(st.session_state.keys())
                    if k.startswith(('res_role_', 'res_main_', 'rad_days_', 'del_check_'))]:
            st.session_state.pop(_wk, None)
        u_holidays = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
        df_loaded = generate_schedule(st.session_state.base_date, st.session_state.week_count, u_holidays, st.session_state.master_schedules, st.session_state.off_slots, supplementary_schedules=st.session_state.supplementary_schedules, biweekly_choice=st.session_state.biweekly_choice)
        # CP-SAT 오버라이드 복원 + df에 적용 (시간/날짜) — 안 하면 df와 배정 시간이 어긋나 개인별 뷰에 중복 표시됨
        time_ovr = data.get("task_time_overrides", {})
        date_ovr = data.get("task_date_overrides", {})
        for tid, t_choice in time_ovr.items():
            df_loaded.loc[df_loaded['task_id'] == tid, 'time'] = t_choice
        _wk_kor = ['월', '화', '수', '목', '금', '토', '일']
        for tid, new_date in date_ovr.items():
            df_loaded.loc[df_loaded['task_id'] == tid, 'date'] = new_date
            try:
                _ndt = datetime.strptime(f"{st.session_state.base_date.year}-{new_date}", "%Y-%m-%d").date()
                df_loaded.loc[df_loaded['task_id'] == tid, 'day'] = _wk_kor[_ndt.weekday()]
                df_loaded.loc[df_loaded['task_id'] == tid, 'week'] = (_ndt - st.session_state.base_date).days // 7 + 1
            except Exception:
                pass
        st.session_state.current_df_all = df_loaded
        st.session_state.cpsat_task_time_overrides = time_ovr
        st.session_state.cpsat_task_date_overrides = date_ovr
        st.session_state.cpsat_shifted_tasks = data.get("shifted_tasks", [])
        st.session_state.cpsat_original_df_dates = data.get("original_df_dates", {})
        # 신규 설정 복원 (옛 백업엔 없을 수 있음 → 기본값으로)
        st.session_state.shift_forbid_date = data.get("shift_forbid_date") or None
        st.session_state.excel_pad_week5 = bool(data.get("excel_pad_week5", False))
        for _k in ('excel_pad_week5_ck_1', 'excel_pad_week5_ck_2'):
            st.session_state[_k] = bool(st.session_state.excel_pad_week5)
        # 불러온 데이터와 어긋나지 않도록 캐시된 엑셀 폐기
        for _k in ('_xls_res_sig', '_xls_res_data', '_xls_prof_sig', '_xls_prof_data'):
            st.session_state.pop(_k, None)
        st.session_state.skip_panjung_obs = list(data.get("skip_panjung_obs", []) or [])
        st.session_state.cpsat_skipped_obs = list(data.get("cpsat_skipped_obs", []) or [])
        st.session_state.cpsat_skipped_prev_month_obs = list(data.get("cpsat_skipped_prev_month_obs", []) or [])
        # relax_prev_month_obs는 체크박스 위젯 key와 동일 — 값만 세팅하면 다음 렌더에서 체크 상태 반영
        st.session_state.relax_prev_month_obs = bool(data.get("relax_prev_month_obs", False))
        st.session_state.drop_prev_month_obs = bool(data.get("drop_prev_month_obs", False))
        _llr = data.get("longleave_range")
        if _llr:
            st.session_state.longleave_range = [float(_llr[0]), float(_llr[1])]
            st.session_state.pop("ll_lo", None); st.session_state.pop("ll_hi", None)
        st.session_state.weekly_extra_bogeonso = bool(data.get("weekly_extra_bogeonso", False))
        st.session_state.weekly_extra_rad = bool(data.get("weekly_extra_rad", False))
        _pjr = data.get("panjung_ranges")
        if _pjr:
            st.session_state.panjung_ranges = dict(_pjr)
            for _yr in ("R2", "R1", "R0"):
                st.session_state.pop(f"pj_lo_{_yr}", None)
                st.session_state.pop(f"pj_hi_{_yr}", None)
            st.session_state.pop("pj_week_max", None)
        # 해제된 하드 룰 복구 + 체크박스 위젯 상태 리셋 (로드값으로 재초기화되도록)
        st.session_state.disabled_rules = list(data.get("disabled_rules", []) or [])
        st.session_state.manual_pins = list(data.get("manual_pins", []) or [])
        for _wk in [k for k in list(st.session_state.keys()) if k.startswith('rule_on_')]:
            st.session_state.pop(_wk, None)
        # (격주 주차 선택·명단 위젯 정리는 스케줄 생성 전에 이미 처리함)
        # 위젯 상태 리셋 (다음 렌더에서 새 skip_panjung_obs로부터 재초기화되도록)
        st.session_state.pop('skip_panjung_obs_multiselect', None)
        # === 스케줄 배정 탭 솔버 설정 복구 (옛 백업엔 없음 → 현재값 유지) ===
        for _k, _cast in [
            ("cpsat_time_limit", int), ("cpsat_manual_multiplier", str),
            ("cpsat_max_broken_pairs", int), ("cpsat_extra_shift", int),
            ("cpsat_prof_repeat_enabled", bool), ("cpsat_prof_repeat_multiplier", float),
            ("cpsat_prof_repeat_slack", int),
        ]:
            if _k in data and data[_k] is not None:
                try:
                    st.session_state[_k] = _cast(data[_k])
                except (TypeError, ValueError):
                    pass
        # 로딩 범위/부등호 (JSON 문자열 키 → int 키로 복원) + 관련 위젯 상태 리셋
        _lr = data.get("loading_ranges")
        if _lr:
            try:
                # 기본값·위젯이 list를 쓰므로 list로 통일 (타입 섞이면 비교/편집에서 오작동)
                st.session_state.loading_ranges = {int(g): list(v) for g, v in _lr.items()}
            except (TypeError, ValueError):
                pass
        _ops = data.get("h17_ops")
        if _ops:
            try:
                st.session_state.h17_ops = {int(g): v for g, v in _ops.items()}
            except (TypeError, ValueError):
                pass
        for _g in range(6):
            for _pfx in ('lr_lo_', 'lr_hi_', 'h17_op_'):
                st.session_state.pop(f"{_pfx}{_g}", None)
        # 배정 후 요약 복원 (없으면 빈 값/제거 — 옛 백업 호환)
        st.session_state.alloc_report = data.get("alloc_report", "")
        for _key, _default in [
            ("prof_repeat_summary", None),
            ("prof_repeat_pairs_data", None),
            ("prof_repeat_by_resident_data", None),
        ]:
            _v = data.get(_key, _default)
            if _v is None:
                st.session_state.pop(_key, None)
            else:
                st.session_state[_key] = _v
        # 연보(보건소 등 강제 배정)는 백업에 없으므로 입력으로부터 재계산해 복원 (개인별/주차별 뷰의 연보 표시용)
        try:
            from cpsat_solver import build_problem_data as _bpd
            _rad = {r['이름']: r.get('영상파견요일', []) for r in st.session_state.residents if "본원 영상" in r.get('역할', [])}
            _pdata = _bpd(df_loaded, st.session_state.residents, st.session_state.resident_leaves,
                          st.session_state.week_count, st.session_state.base_date, u_holidays,
                          bogeonso_substitutes=st.session_state.bogeonso_substitutes, rad_days=_rad,
                          student_practices=st.session_state.student_practices)
            st.session_state.cpsat_forced_assignments = _pdata.get('forced_assignments', {})
        except Exception:
            st.session_state.cpsat_forced_assignments = {}
        # 개인별/주차별 뷰용 res_daily_slots 재구성 (함수 정의 이후 rerun에서 처리)
        st.session_state._needs_slot_rebuild = True
        st.rerun()

if st.session_state.current_df_all.empty:
    user_holidays = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
    st.session_state.current_df_all = generate_schedule(st.session_state.base_date, st.session_state.week_count, user_holidays, st.session_state.master_schedules, st.session_state.off_slots, supplementary_schedules=st.session_state.supplementary_schedules, biweekly_choice=st.session_state.get('biweekly_choice'))

df_all = st.session_state.current_df_all

def build_res_daily_slots(assignments, forced_assignments=None, df=None):
    """assignments(+공휴일/휴가/메인외래/학생실습/연보/영상 오버레이)로 전공의별 daily_slots 재구성.
    CP-SAT 자동배정과 '수동 배정 저장' 모두 이 함수를 써서 전공의 개인별/주차별 현황 뷰가
    항상 최신 assignments와 일치하도록 한다.
    df: 시간/날짜 오버라이드가 적용된 task df. None이면 st.session_state.current_df_all 사용.
        (CP-SAT 직후엔 current_df_all이 아직 갱신 전이므로 반드시 df_gen을 넘겨야 함)
    """
    if forced_assignments is None:
        forced_assignments = {}
    u_hols = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
    cur_df = df if df is not None else st.session_state.current_df_all
    rad_days_d = {r['이름']: r.get('영상파견요일', []) for r in st.session_state.residents if "본원 영상" in r.get('역할', [])}
    wd_to_idx = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4}
    res_slots = {}
    for r in st.session_state.residents:
        name = r['이름']
        daily = {}
        for w in range(st.session_state.week_count):
            for d_idx in range(5):
                ds = (st.session_state.base_date + timedelta(days=w*7+d_idx)).strftime("%m-%d")
                daily[ds] = {'오전': None, '오후': None}
        # 공휴일 (최우선)
        for ds in daily:
            if ds in u_hols:
                daily[ds]['오전'] = '공휴일'; daily[ds]['오후'] = '공휴일'
        # 휴가
        for l in st.session_state.resident_leaves:
            if l['이름'] == name and l['날짜'] in daily:
                if daily[l['날짜']]['오전'] != '공휴일': daily[l['날짜']]['오전'] = l['종류']
                if daily[l['날짜']]['오후'] != '공휴일': daily[l['날짜']]['오후'] = l['종류']
        # 메인외래
        main = r.get('메인외래', '선택안함')
        if main != '선택안함' and main in wd_to_idx:
            d_idx = wd_to_idx[main]
            for w in range(st.session_state.week_count):
                ds = (st.session_state.base_date + timedelta(days=w*7+d_idx)).strftime("%m-%d")
                if ds in daily and daily[ds]['오전'] is None: daily[ds]['오전'] = '메인외래'
                if ds in daily and daily[ds]['오후'] is None: daily[ds]['오후'] = '메인외래'
        # 학생실습
        for sp in st.session_state.student_practices:
            if sp['이름'] == name and sp['날짜'] in daily and daily[sp['날짜']][sp['시간']] is None:
                daily[sp['날짜']][sp['시간']] = '학생실습'
        # 연보 (forced_assignments)
        for (fname, fdate, ftime), flabel in forced_assignments.items():
            if fname == name and fdate in daily and daily[fdate][ftime] is None:
                daily[fdate][ftime] = flabel
        # 영상 파견
        for rd in rad_days_d.get(name, []):
            if rd in wd_to_idx:
                d_idx = wd_to_idx[rd]
                for w in range(st.session_state.week_count):
                    ds = (st.session_state.base_date + timedelta(days=w*7+d_idx)).strftime("%m-%d")
                    if ds in daily:
                        if daily[ds]['오전'] is None: daily[ds]['오전'] = '영상'
                        if daily[ds]['오후'] is None: daily[ds]['오후'] = '영상'
        # 배정 결과
        for tid, assignee in assignments.items():
            if assignee != name: continue
            tr = cur_df[cur_df['task_id'] == tid]
            if tr.empty: continue
            row = tr.iloc[0]; ds, ttime = row['date'], row['time']
            if ds in daily and daily[ds].get(ttime) is None:
                daily[ds][ttime] = tid
        res_slots[name] = {'daily_slots': daily}
    return res_slots

# 백업 로드 직후: assignments 기준으로 res_daily_slots 재구성 (개인별/주차별 뷰 동기화)
if st.session_state.pop('_needs_slot_rebuild', False):
    st.session_state.res_daily_slots = build_res_daily_slots(
        st.session_state.assignments,
        st.session_state.get('cpsat_forced_assignments', {})
    )

st.markdown("<div style='font-size: 1.1rem; font-weight: 600; margin-bottom: 15px; color: #555555; text-align: left;'>👨‍💻 Made by 45기 변준혁 문의 T. 010-4937-1111</div>", unsafe_allow_html=True)

# --- 메인 탭 (순서 변경: 사용법이 가장 왼쪽) ---
tabs = st.tabs(["📖 사용법", "👨‍🏫 교수별 시간표", "📊 주차별 가이드", "⚙️ 규칙 설정", "👨‍⚕️ 전공의 명단", "🌴 전공의 휴가", "📝 스케줄 배정", "📅 전공의 개인별", "🗓️ 주차별 전체 현황(Excel양식)", "✅ 스케줄 검증"])

# 탭 인덱스 0: 사용법
with tabs[0]:
    show_manual()

# 탭 인덱스 4 (기존 3): 전공의 명단
with tabs[4]:
    st.subheader("👨‍⚕️ 전공의 명단 및 역할 관리")
    st.caption(
        "💡 연차를 고른 뒤 이름 여러 개를 **공백·쉼표·줄바꿈**으로 한 번에 입력하고 [➕ 일괄 등록]. "
        "역할·메인외래·영상 파견 요일은 등록 후 아래 카드에서 자유롭게 바꾼 다음 "
        "맨 아래 **[💾 모든 변경사항 적용]** 한 번만 누르세요 (매번 로딩 없음). "
        "삭제는 카드 우측 ✖ 버튼."
    )
    # ── 일괄 등록 폼 (submit 후 자동 클리어) ────────────────────────────
    with st.form("res_bulk_add", clear_on_submit=True, border=False):
        cols = st.columns([1, 4, 1])
        with cols[0]:
            r_year = st.selectbox("연차", ["R3", "R2", "R1", "R0"])
        with cols[1]:
            r_names_raw = st.text_input(
                "이름 (여러 명 가능 — 공백·쉼표·줄바꿈 구분)",
                placeholder="예: 홍길동 김철수 이영희"
            )
        with cols[2]:
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)  # 라벨 높이만큼 정렬
            submitted = st.form_submit_button("➕ 일괄 등록", use_container_width=True)
    if submitted:
        raw = (r_names_raw or "").strip()
        if raw:
            new_names = [n.strip() for n in re.split(r"[\s,]+", raw) if n.strip()]
            existing = {res['이름'] for res in st.session_state.residents}
            added, skipped = 0, []
            for n in new_names:
                if n in existing:
                    skipped.append(n)
                    continue
                st.session_state.residents.append({
                    "연차": r_year, "이름": n, "역할": [],
                    "메인외래": "선택안함", "영상파견요일": []
                })
                existing.add(n)
                added += 1
            if added:
                st.toast(f"{r_year} {added}명 추가됨", icon="✅")
            if skipped:
                st.toast(f"중복으로 건너뜀: {', '.join(skipped)}", icon="⚠️")
            st.rerun()

    st.markdown("---")
    # ── 인라인 편집 폼 (폼 안에서는 위젯 변경 시 rerun 없음, 적용 누르면 한 번에 commit) ─
    role_options_by_year = {
        "R3": ["본원 영상", "의국수석", "교육수석", "학생수석", "진료수석"],
        "R2": ["연건 보건소"],
        "R1": ["의국 처음"],
        "R0": ["의국 처음"],
    }
    main_options = ["선택안함", "월", "화", "수", "목", "금"]
    rad_day_options = ["월", "화", "수", "목", "금"]

    with st.form("residents_edit_form", border=False):
        c1, c2, c3, c4 = st.columns(4)
        for col_idx, year in enumerate(["R3", "R2", "R1", "R0"]):
            with [c1, c2, c3, c4][col_idx]:
                st.markdown(f"#### 🔹 {year}")
                for r in [res for res in st.session_state.residents if res["연차"] == year]:
                    # 위젯 key는 이름 기반 (이름 중복 불가) — 삭제 후 리스트 재정렬돼도 key가 안 섞임
                    rkey = r['이름']
                    with st.container(border=True):
                        rc1, rc2 = st.columns([3, 2])
                        rc1.markdown(f"**{r['이름']}**")
                        # 🗑 삭제 표시 체크박스 (적용 시 체크된 사람 일괄 삭제)
                        rc2.checkbox(
                            "🗑 삭제",
                            key=f"del_check_{rkey}",
                            value=False,
                            help=f"체크 후 [💾 적용]을 누르면 {r['이름']} 삭제",
                        )
                        # 역할 (위젯만 그리기, 적용은 폼 submit 시)
                        if role_options_by_year[year]:
                            st.multiselect(
                                "역할",
                                options=role_options_by_year[year],
                                default=r.get('역할', []),
                                key=f"res_role_{rkey}",
                                placeholder="역할 선택",
                                label_visibility="collapsed",
                            )
                        # 메인외래 (R3만)
                        if year == "R3":
                            cur_main = r.get('메인외래', "선택안함")
                            st.selectbox(
                                "메인외래",
                                options=main_options,
                                index=main_options.index(cur_main) if cur_main in main_options else 0,
                                key=f"res_main_{rkey}",
                                label_visibility="collapsed",
                            )
                        # 영상 파견 요일 (현재 저장된 역할 기준 — 폼 안에서는 실시간 토글 안 됨)
                        if "본원 영상" in r.get('역할', []):
                            st.multiselect(
                                "영상 파견 요일 (매주)",
                                options=rad_day_options,
                                default=r.get('영상파견요일', []),
                                key=f"rad_days_{rkey}",
                                placeholder="영상 파견 요일",
                                label_visibility="collapsed",
                                help="선택한 요일에는 매주 오전+오후 모두 '영상'으로 고정됩니다."
                            )
                        # 현재 저장값 요약 (적용 전이라 폼 안의 미반영 변경은 안 보임)
                        bits = []
                        if r.get('역할'):
                            bits.append(f"<small style='color:#1f77b4;'>{', '.join(r['역할'])}</small>")
                        if r.get('메인외래', "선택안함") != "선택안함":
                            bits.append(f"<small style='color:#d62728;'>메인: {r['메인외래']}</small>")
                        rad_days_now = r.get('영상파견요일', [])
                        if rad_days_now:
                            bits.append(f"<small style='color:#27ae60;'>📡 영상: {', '.join(rad_days_now)}</small>")
                        if bits:
                            st.markdown(" | ".join(bits), unsafe_allow_html=True)

        st.markdown("---")
        st.caption(
            "💡 '본원 영상' 역할을 새로 추가했다면 [💾 적용] 한 번 누른 후 다시 와야 영상 파견 요일 위젯이 나타납니다."
        )
        submitted_apply = st.form_submit_button(
            "💾 모든 변경사항 적용", type="primary", use_container_width=True
        )

    # === 폼 처리 (폼 밖) ===
    if submitted_apply:
        # 1) 체크된 사람들 일괄 삭제 (위젯 key = 이름 기반이라 인덱스 재정렬 무관)
        deleted_names = [r['이름'] for r in st.session_state.residents
                         if st.session_state.get(f"del_check_{r['이름']}", False)]
        if deleted_names:
            st.session_state.residents = [r for r in st.session_state.residents
                                          if r['이름'] not in deleted_names]
        # 삭제된 사람의 widget key 정리
        for nm in deleted_names:
            for prefix in ('del_check_', 'res_role_', 'res_main_', 'rad_days_'):
                st.session_state.pop(f"{prefix}{nm}", None)

        # 2) 남은 사람들의 변경 commit (key = 이름 기반)
        changed_count = 0
        for r in st.session_state.residents:
            rkey = r['이름']
            # 역할
            if role_options_by_year.get(r['연차']):
                wkey = f"res_role_{rkey}"
                if wkey in st.session_state:
                    new_v = list(st.session_state[wkey])
                    if new_v != list(r.get('역할', [])):
                        r['역할'] = new_v
                        changed_count += 1
            # 메인외래 (R3)
            if r['연차'] == "R3":
                wkey = f"res_main_{rkey}"
                if wkey in st.session_state:
                    new_v = st.session_state[wkey]
                    if new_v != r.get('메인외래', "선택안함"):
                        r['메인외래'] = new_v
                        changed_count += 1
            # 영상 파견 요일 — 적용 후 r['역할']에 본원 영상이 있는 경우만
            if "본원 영상" in r.get('역할', []):
                wkey = f"rad_days_{rkey}"
                if wkey in st.session_state:
                    new_v = list(st.session_state[wkey])
                    if new_v != list(r.get('영상파견요일', [])):
                        r['영상파견요일'] = new_v
                        changed_count += 1
            else:
                # 본원 영상 빠진 경우 — 영상파견요일 자동 비우기 (저장된 게 있을 때만)
                if r.get('영상파견요일'):
                    r['영상파견요일'] = []
                    changed_count += 1
        msg_parts = []
        if deleted_names:
            msg_parts.append(f"🗑 {len(deleted_names)}명 삭제 ({', '.join(deleted_names)})")
        if changed_count:
            msg_parts.append(f"✅ {changed_count}건 변경 적용")
        if msg_parts:
            st.toast(" · ".join(msg_parts), icon="✅")
        else:
            st.toast("변경 사항 없음", icon="ℹ️")
        st.rerun()

    # ── 판정참관 배정 제외 대상 (건증 판정은 받되, 그 쌍의 판정 참관은 미배정) ──
    st.markdown("---")
    try:
        with st.container(border=True):
            st.markdown("#### 🚫 건증 판정참관 배정 제외 대상")
            st.caption(
                "선택된 전공의에게 건증 판정이 배정될 경우, 그 쌍의 **판정 참관은 어디에도 배정되지 않습니다** "
                "(판정 자체는 정상 배정, 참관 슬롯은 비어서 다른 task가 그 자리에 채워질 수 있음)."
            )
            _all_names = [r['이름'] for r in st.session_state.residents]
            _skey = 'skip_panjung_obs_multiselect'
            # 위젯 상태를 현재 residents 목록에 맞춰 정리 (없는 이름 자동 제거)
            if _skey in st.session_state:
                _cleaned = [n for n in st.session_state[_skey] if n in _all_names]
                if _cleaned != list(st.session_state[_skey]):
                    st.session_state[_skey] = _cleaned
            else:
                st.session_state[_skey] = [n for n in st.session_state.get('skip_panjung_obs', []) if n in _all_names]
            st.multiselect(
                "제외 대상 전공의",
                options=_all_names,
                key=_skey,
                placeholder="이름 선택 (복수 가능)",
                label_visibility="collapsed",
            )
            st.session_state.skip_panjung_obs = list(st.session_state[_skey])

            st.markdown("---")
            # 이전 달에 생성된 참관(orphan = 짝 차리/판정이 이번 달 밖) 처리 옵션
            st.checkbox(
                "이전 달에 생성된 참관은 배정 안 해도 됨 (미배정 페널티를 아주 작게)",
                key="relax_prev_month_obs",
                help="체크 시: 짝(차리/판정)이 지난달이라 이번 달엔 참관만 남은 task는 '가능하면 배정, "
                     "아니면 미배정'으로 둡니다(강제 배정 안 함). 해제 시: 기존처럼 강제로 배정합니다. "
                     "참관 미배정 대상(위 목록)에게는 이 참관을 절대 배정하지 않습니다(체크와 무관).",
            )
            st.checkbox(
                "이전 달에 생성된 참관은 **아예 배정하지 않음** (전부 미배정 처리)",
                key="drop_prev_month_obs",
                help="체크 시: 짝(차리/판정)이 지난달이라 참관만 남은 task를 누구에게도 배정하지 않고 "
                     "전부 미배정으로 둡니다. 위 '배정 안 해도 됨'보다 우선합니다. "
                     "⚠️ 그만큼 배정할 일이 줄어 로딩 하한(H15)을 못 채우면 해가 없을 수 있습니다.",
            )
            if st.session_state.get("drop_prev_month_obs"):
                st.warning("⚠️ 이전 달 참관을 **전부 미배정**합니다. 해가 없으면(INFEASIBLE) 이 체크를 풀거나 로딩 하한을 낮춰보세요.")
            st.caption("💡 미배정된 이전 달 참관은 '검증' 탭에서 별도로 집계됩니다(진짜 미배정 아님).")
    except Exception as _e:
        st.warning(f"판정참관 제외 위젯 오류(무시하고 계속): {_e}")

# 탭 인덱스 5 (기존 4): 전공의 휴가
with tabs[5]:
    st.subheader("🌴 전공의 휴가 입력")
    res_names = [r["이름"] for r in st.session_state.residents]
    with st.form("leave_form", clear_on_submit=True):
        cols = st.columns([2, 3, 2, 1])
        l_name, l_dates = cols[0].selectbox("전공의", res_names if res_names else ["없음"]), cols[1].multiselect("날짜", options=available_dates)
        l_type = cols[2].selectbox("종류", ["직전휴가", "사전휴가", "Off"])
        if cols[3].form_submit_button("등록") and res_names:
            for d_str in l_dates: st.session_state.resident_leaves.append({"이름": l_name, "날짜": d_str.split(" ")[0], "종류": l_type})
            st.rerun()
    if st.session_state.resident_leaves:
        with st.expander(f"📋 등록된 휴가 ({len(st.session_state.resident_leaves)}건)", expanded=True):
            # 신청자(전공의)별로 묶어서 표시 — 원래 인덱스는 삭제용으로 보존
            from collections import defaultdict as _dd
            groups = _dd(list)
            for i, lv in enumerate(st.session_state.resident_leaves):
                groups[lv['이름']].append((i, lv))
            # 명단 순서 우선, 명단에 없는 이름은 뒤에 이름순
            ordered_names = [n for n in res_names if n in groups] + sorted(n for n in groups if n not in res_names)
            for name in ordered_names:
                items = sorted(groups[name], key=lambda x: x[1]['날짜'])
                st.markdown(f"**👤 {name}** · {len(items)}건")
                n_cols = 3
                grid_cols = st.columns(n_cols)
                for pos, (i, lv) in enumerate(items):
                    with grid_cols[pos % n_cols]:
                        cc1, cc2 = st.columns([5, 1])
                        cc1.write(f"{lv['날짜']} [{lv['종류']}]")
                        if cc2.button("X", key=f"leave_del_{i}"):
                            st.session_state.resident_leaves.pop(i)
                            st.rerun()

# 탭 인덱스 6 (기존 5): 스케줄 배정
with tabs[6]:
    st.subheader("📝 스케줄 배정 및 자동 생성")
    res_names = [r["이름"] for r in st.session_state.residents]
    
    # === [신규] 연건 보건소 휴가 대체자 설정 ===
    st.markdown("#### 🏥 연건 보건소 휴가 대체자 설정")
    # 연건 보건소 담당자 찾기
    bogeonso_residents = [r["이름"] for r in st.session_state.residents if "연건 보건소" in r.get("역할", [])]
    # 연건 보건소 담당자가 휴가인 날짜 추출
    bogeonso_leave_dates = sorted(set(
        l["날짜"] for l in st.session_state.resident_leaves if l["이름"] in bogeonso_residents
    ))
    
    if not bogeonso_residents:
        st.info("ℹ️ 연건 보건소 역할이 지정된 전공의가 없습니다. (전공의 명단 탭에서 R2에게 '연건 보건소' 역할을 부여하세요)")
    elif not bogeonso_leave_dates:
        st.info(f"ℹ️ 현재 연건 보건소 담당자({', '.join(bogeonso_residents)})의 휴가가 등록되어 있지 않습니다. 휴가가 등록되면 이 영역에서 대체자를 지정할 수 있습니다.")
    else:
        st.caption(f"💡 연건 보건소 담당자: **{', '.join(bogeonso_residents)}** — 휴가 등록된 날짜의 연보 일정을 누가 대신 갈지 지정하세요. (복수 선택 가능, 여러 명 선택 시 슬롯이 라운드로빈 분배됩니다)")
        st.caption("⚠️ 여기서 설정한 대체자는 자동 스케줄 생성 시 **무조건** 해당 날짜의 연보로 배정됩니다. 그리고 그 외 비어있는 시간에는 다른 일반 업무도 추가 배정될 수 있습니다.")
        
        # 휴가 날짜 중 더 이상 유효하지 않은 항목 정리
        st.session_state.bogeonso_substitutes = {
            d: subs for d, subs in st.session_state.bogeonso_substitutes.items() if d in bogeonso_leave_dates
        }
        
        # 대체자 후보 (연건 보건소 본인 제외)
        sub_candidates = [n for n in res_names if n not in bogeonso_residents]
        
        for d_str in bogeonso_leave_dates:
            # 해당 날짜에 휴가 낸 사람과 종류 표시
            leave_info = next((l for l in st.session_state.resident_leaves if l["날짜"] == d_str and l["이름"] in bogeonso_residents), None)
            label_extra = ""
            if leave_info:
                label_extra = f" — {leave_info['이름']} [{leave_info['종류']}]"
            
            # 요일 계산
            try:
                year = st.session_state.base_date.year
                dt = datetime.strptime(f"{year}-{d_str}", "%Y-%m-%d")
                day_name = ["월", "화", "수", "목", "금", "토", "일"][dt.weekday()]
                slot_info = "오전+오후" if dt.weekday() in [1, 3] else "오전만"
            except:
                day_name = "?"
                slot_info = ""
            
            current_subs = st.session_state.bogeonso_substitutes.get(d_str, [])
            # 유효하지 않은 이전 선택 정리
            current_subs = [s for s in current_subs if s in sub_candidates]
            
            selected = st.multiselect(
                f"📅 **{d_str} ({day_name})** {label_extra} — 연보 {slot_info}",
                options=sub_candidates,
                default=current_subs,
                key=f"bogeonso_sub_{d_str}",
                help="여러 명 선택 시 라운드로빈으로 슬롯을 나눠가져갑니다."
            )
            if selected:
                st.session_state.bogeonso_substitutes[d_str] = selected
            elif d_str in st.session_state.bogeonso_substitutes:
                del st.session_state.bogeonso_substitutes[d_str]
    
    st.markdown("---")
    
    st.markdown("#### 🎓 학생 실습 선입력 (로딩 +1)")
    with st.form("student_form", clear_on_submit=True):
        sc1, sc2, sc3 = st.columns(3)
        s_name = sc1.selectbox("대상자", res_names if res_names else ["없음"])
        s_date = sc2.selectbox("날짜", get_date_options(st.session_state.base_date, st.session_state.week_count))
        s_time = sc3.selectbox("시간", ["오전", "오후"])
        if st.form_submit_button("실습 등록") and res_names:
            st.session_state.student_practices.append({"이름": s_name, "날짜": s_date.split(" ")[0], "시간": s_time})
            st.rerun()
    for i, sp in enumerate(st.session_state.student_practices):
        c1, c2, c3 = st.columns([3, 1, 1])
        c1.write(f"· **{sp['이름']}** - {sp['날짜']} {sp['시간']} 실습")
        if c2.button("삭제", key=f"sp_{i}"):
            st.session_state.student_practices.pop(i); st.rerun()
    
    st.markdown("---")
    st.markdown("#### 💉 통증클리닉 특수 배정")
    st.session_state.pain_applicants = st.multiselect(
        "박진호 통증클리닉 사전 신청자 (차리/참관 세트 최우선 배정)",
        options=res_names,
        default=[n for n in st.session_state.pain_applicants if n in res_names]
    )
    
    st.markdown("---")
    # === [신규] 사전 진단 박스 ===
    st.markdown("#### 📊 사전 진단")
    if not st.session_state.residents:
        st.caption("전공의 명단을 먼저 등록해주세요.")
        diagnosis_result = None
    else:
        try:
            u_holidays_for_diag = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
            rad_days_for_diag = {
                r['이름']: r.get('영상파견요일', [])
                for r in st.session_state.residents
                if "본원 영상" in r.get('역할', [])
            }
            from utils import pre_assignment_diagnosis
            diagnosis_result = pre_assignment_diagnosis(
                st.session_state.residents,
                st.session_state.resident_leaves,
                st.session_state.week_count,
                st.session_state.base_date,
                u_holidays_for_diag,
                st.session_state.off_slots,
                st.session_state.master_schedules,
                supplementary_schedules=st.session_state.supplementary_schedules,
                rad_days=rad_days_for_diag,
                student_practices=st.session_state.student_practices,
                bogeonso_substitutes=st.session_state.bogeonso_substitutes,
                loading_ranges=st.session_state.loading_ranges,
                skip_panjung_obs=st.session_state.get('skip_panjung_obs') or None,
                biweekly_choice=st.session_state.get('biweekly_choice'),
            )
            if diagnosis_result['status'] == "적합":
                st.markdown(f"<div style='padding:10px; background:#E8F8F5; border-left:4px solid #1abc9c; border-radius:3px;'>{diagnosis_result['message']}</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='padding:10px; background:#FADBD8; border-left:4px solid #e74c3c; border-radius:3px;'>{diagnosis_result['message']}</div>", unsafe_allow_html=True)
            st.caption(f"💡 사용자 지정 평균 로딩 {diagnosis_result['ideal_avg_target_mult']:.2f} | 실제 필요 평균 로딩 {diagnosis_result['required_avg_loading']:.2f} | 자동 상향 배수 {diagnosis_result['multiplier']:.2f}")
        except Exception as e:
            st.error(f"진단 중 오류: {e}")
            diagnosis_result = None

    st.markdown("---")
    # === CP-SAT 솔버 설정 (CP-SAT 모드 상시 사용) ===
    st.session_state.use_cpsat = True
    st.caption("🧮 CP-SAT 솔버로 배정합니다 (모든 룰 보장, 해가 없으면 알림).")
    mode_col2, mode_col3 = st.columns(2)
    with mode_col2:
        st.session_state.cpsat_time_limit = st.number_input(
            "제한시간(초)", min_value=10, max_value=600, value=st.session_state.cpsat_time_limit, step=10
        )
    with mode_col3:
        # 배율 옵션: 자동 또는 1.025~1.300 (0.025 단위)
        mult_options = ["자동"] + [f"{x/1000:.3f}" for x in range(1025, 1301, 25)]
        current = st.session_state.cpsat_manual_multiplier
        if current not in mult_options:
            current = "자동"
        st.session_state.cpsat_manual_multiplier = st.selectbox(
            "배율 강제 지정",
            options=mult_options,
            index=mult_options.index(current),
            help="'자동'이면 시스템이 사전 진단으로 산출. 직접 지정하면 그 배율로 풀이."
        )
    set_col1, set_col2 = st.columns(2)
    with set_col1:
        st.session_state.cpsat_max_broken_pairs = st.number_input(
            "깨도 되는 pairing 최대 개수", min_value=0, max_value=50,
            value=st.session_state.cpsat_max_broken_pairs, step=1,
            help="이 개수까지 차리/판정+참관 묶음을 분리 허용. (건증·박진호·권혁태 묶음은 항상 보호되어 이 한도와 무관하게 안 깨짐)"
        )
    with set_col2:
        st.session_state.cpsat_extra_shift = st.number_input(
            "차리 추가 -1 이동 허용 개수", min_value=0, max_value=100,
            value=st.session_state.cpsat_extra_shift, step=1,
            help="슬롯 부족 날짜는 자동으로 차리/판정 -1 이동이 허용됩니다. 그 외에도 이 개수만큼 '차리'(판정 제외)를 직전 평일로 이동 허용 (빡빡해서 해가 없을 때 늘려보세요)."
        )

    # === 🔧 하드 룰 임시 해제 (기본 접힘 · 전부 체크 상태) ===
    with st.expander("🔧 하드 룰 임시 해제 (고급) — 기본값: 전부 적용", expanded=False):
        st.caption(
            "체크를 **해제한 룰은 이번 배정에서 적용하지 않습니다**. "
            "해가 없거나(INFEASIBLE) 특정 칸을 예외적으로 채워야 할 때만 잠깐 끄세요. "
            "예: **H10을 끄면 조비룡·박민선 외래와 예진에 R2도 배정**할 수 있습니다. "
            "⚠️ H1/H2(한 슬롯 1개)·H5(휴가·공휴일)는 기본 전제라 끌 수 없습니다."
        )
        _rc1, _rc2 = st.columns(2)
        _new_disabled = []
        for _i, (_rkey, _rlabel, _rhelp) in enumerate(DISABLEABLE_RULES):
            with (_rc1 if _i % 2 == 0 else _rc2):
                _checked = st.checkbox(
                    _rlabel,
                    value=(_rkey not in st.session_state.disabled_rules),
                    key=f"rule_on_{_rkey}",
                    help=_rhelp,
                )
                if not _checked:
                    _new_disabled.append(_rkey)
        st.session_state.disabled_rules = _new_disabled
        if _new_disabled:
            _labels = [l for k, l, _ in DISABLEABLE_RULES if k in _new_disabled]
            st.warning(f"⚠️ 해제된 룰 {len(_new_disabled)}개 — {', '.join(_labels)}")
        else:
            st.success("✅ 모든 하드 룰 적용 중 (기본 상태)")

    # === 📌 직접 지정 배정 (부족한 날 예외 처리) ===
    with st.expander("📌 직접 지정 배정 — 부족한 날 자동 확인 + 예외 배정", expanded=False):
        st.caption(
            "**조비룡·박민선 외래참관/예진**은 원래 R3만 받습니다(H10). "
            "그 시간에 가능한 R3가 모자란 날을 자동으로 찾아 알려주고, "
            "그 칸만 **R2 등 다른 전공의로 직접 지정**할 수 있습니다. "
            "지정한 칸은 H10을 우회해 **그 사람에게 강제 배정**됩니다. (H10 체크는 켜둔 채로 사용하세요)"
        )
        # (1) 부족한 날 자동 검증
        try:
            from cpsat_solver import find_r3_only_shortage
            _u_hol = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
            _rad_sh = {r['이름']: r.get('영상파견요일', []) for r in st.session_state.residents
                       if "본원 영상" in r.get('역할', [])}
            _short = find_r3_only_shortage(
                st.session_state.current_df_all, st.session_state.residents,
                st.session_state.resident_leaves, st.session_state.week_count,
                st.session_state.base_date, _u_hol,
                bogeonso_substitutes=st.session_state.bogeonso_substitutes,
                rad_days=_rad_sh, student_practices=st.session_state.student_practices,
            )
        except Exception as _e:
            _short = []
            st.warning(f"부족 날짜 검증 실패(무시하고 계속): {_e}")
        _wd_kor = ['월', '화', '수', '목', '금', '토', '일']
        if _short:
            for _s in _short:
                try:
                    _dt = datetime.strptime(f"{st.session_state.base_date.year}-{_s['date']}", "%Y-%m-%d").date()
                    _dn = _wd_kor[_dt.weekday()]
                except Exception:
                    _dn = '?'
                st.error(
                    f"🔴 **{_s['date']}({_dn}) {_s['time']} — {_s['shortage']}개 부족**  \n"
                    f"업무: {', '.join(_s['tasks'])}  \n"
                    f"가능한 R3: {', '.join(_s['available_r3']) or '없음'}  \n"
                    f"★ 지정 가능: {', '.join(_s['candidates_r2']) or '없음'}"
                )
        else:
            st.success("✅ R3 전용(예진·조비룡/박민선 외래참관) 자리가 부족한 날이 없습니다.")

        # (2) 지정 추가
        _df_now = st.session_state.current_df_all
        _date_opts = sorted(_df_now['date'].unique().tolist()) if not _df_now.empty else []
        # 부족한 날을 기본 선택으로
        _def_idx = 0
        if _short and _short[0]['date'] in _date_opts:
            _def_idx = _date_opts.index(_short[0]['date'])
        pc1, pc2, pc3, pc4 = st.columns([1.2, 2.5, 1.5, 1])
        with pc1:
            _pin_date = st.selectbox("날짜", _date_opts, index=_def_idx, key="pin_date") if _date_opts else None
        with pc2:
            _task_opts = sorted(_df_now[_df_now['date'] == _pin_date]['task'].unique().tolist()) if _pin_date else []
            _pin_task = st.selectbox("업무", _task_opts, key="pin_task") if _task_opts else None
        with pc3:
            _name_opts = [r['이름'] for r in st.session_state.residents]
            _pin_person = st.selectbox("담당자", _name_opts, key="pin_person") if _name_opts else None
        with pc4:
            st.markdown("<div style='height:28px;'></div>", unsafe_allow_html=True)
            if st.button("➕ 지정", use_container_width=True, key="pin_add"):
                if _pin_date and _pin_task and _pin_person:
                    _exists = any(p['date'] == _pin_date and p['task'] == _pin_task
                                  for p in st.session_state.manual_pins)
                    if _exists:
                        st.warning("이미 그 날짜·업무에 지정이 있습니다. 먼저 삭제하세요.")
                    else:
                        st.session_state.manual_pins.append(
                            {"date": _pin_date, "task": _pin_task, "person": _pin_person})
                        st.rerun()

        # (3) 등록된 지정 목록
        if st.session_state.manual_pins:
            st.markdown(f"**등록된 지정 ({len(st.session_state.manual_pins)}건)**")
            for _i, _p in enumerate(list(st.session_state.manual_pins)):
                _c1, _c2 = st.columns([6, 1])
                _c1.write(f"· **{_p['date']}** · {_p['task']} → **{_p['person']}**")
                if _c2.button("X", key=f"pin_del_{_i}"):
                    st.session_state.manual_pins.pop(_i)
                    st.rerun()
            st.caption("⚠️ 지정한 사람이 그날 휴가·메인외래 등으로 막혀 있으면 해가 없습니다(INFEASIBLE). 위 '지정 가능' 명단에서 고르세요.")

    # === ⛔ Shift 금지 날짜 ===
    with st.container(border=True):
        st.markdown("##### ⛔ Shift 금지 날짜")
        st.caption(
            "차리/판정이 **-1일(직전 평일)로 이동해서 이 날짜에 배치되는 것**을 방지합니다. "
            "예: 실제 근무 종료일 다음날(사전휴가 시작일)을 지정하면, 그 날로 shift되어 미배정 세션이 생기는 걸 막습니다."
        )
        sf_col1, sf_col2 = st.columns([1, 3])
        with sf_col1:
            _sf_enabled = st.checkbox(
                "사용",
                value=bool(st.session_state.get('shift_forbid_date')),
                key='shift_forbid_enabled',
            )
        with sf_col2:
            _default_sf = None
            _cur_sf = st.session_state.get('shift_forbid_date')
            if _cur_sf:
                try:
                    _default_sf = datetime.strptime(
                        f"{st.session_state.base_date.year}-{_cur_sf}", "%Y-%m-%d"
                    ).date()
                except Exception:
                    _default_sf = None
            if _sf_enabled:
                _picked = st.date_input(
                    "shift 금지 날짜",
                    value=_default_sf if _default_sf else st.session_state.base_date,
                    key='shift_forbid_date_input',
                )
                st.session_state.shift_forbid_date = _picked.strftime("%m-%d")
            else:
                st.session_state.shift_forbid_date = None
                st.caption("체크 해제 시 shift 금지 없음 (기본)")

    # === 🎯 교수 중복 최소화 (S5) ε-constraint ===
    with st.container(border=True):
        st.markdown("##### 🎯 교수 중복 최소화 (S5)")
        st.caption(
            "같은 교수+카테고리(외래·암외래 통합 / 건증 / 통증클리닉 등)에서 "
            "한 전공의가 **4회 이상** 받는 페어를 카운트. "
            "점수 = `max(0, count-3)²` → **4회=1 · 5회=4 · 6회=9 · 7회=16**."
        )
        pc1, pc2 = st.columns([1, 1])
        with pc1:
            st.session_state.cpsat_prof_repeat_enabled = st.checkbox(
                "강화 모드 (ε-constraint) 사용",
                value=st.session_state.cpsat_prof_repeat_enabled,
                help=(
                    "ON: 2-phase 솔브. 1차로 실제 도달 가능한 최소값 L*를 계산하고, "
                    "2차에서 중복지수 ≤ ⌈배수 × L*⌉ hard 제약 추가 후 본 솔브. "
                    "솔브 시간이 최대 약 2배까지 늘 수 있습니다.\n"
                    "OFF: 기존 단일 솔브 (소프트 페널티 가중치 1)."
                ),
            )
        with pc2:
            st.session_state.cpsat_prof_repeat_multiplier = st.slider(
                "허용 배수 (L* 대비)",
                min_value=1.00, max_value=2.00,
                value=float(st.session_state.cpsat_prof_repeat_multiplier),
                step=0.05,
                disabled=not st.session_state.cpsat_prof_repeat_enabled,
                help="1.00 = 진짜 최소만 허용 / 1.20 = 1.2배까지 허용 / 2.00 = 거의 안 봄",
            )
            st.session_state.cpsat_prof_repeat_slack = st.number_input(
                "최소 여유 점수 (L+slack)",
                min_value=0, max_value=50,
                value=int(st.session_state.cpsat_prof_repeat_slack),
                step=1,
                disabled=not st.session_state.cpsat_prof_repeat_enabled,
                help=(
                    "cap = max(⌈배수×L⌉, **L + slack**). "
                    "L이 작아 배수만으로는 여유가 거의 없을 때 절대 여유 보장. "
                    "기본 5점 (= 5회 페어 1개 또는 4회 페어 5개 분량). "
                    "0이면 비활성 (예전 배수만 사용)."
                ),
            )

        # === 즉시 계산 가능한 이론적 점수 lower bound (매우 낙관적) ===
        try:
            # 일반 전공의 (보건소/영상 제외)
            _gen_residents = []
            for _r in st.session_state.residents:
                _roles = _r.get('역할', [])
                if "연건 보건소" in _roles: continue
                if "본원 영상" in _roles and _r.get('영상파견요일'): continue
                _gen_residents.append(_r)
            K_gen = max(1, len(_gen_residents))

            # 교수별 주간 task 발생량 (master_schedules 기준)
            ms = st.session_state.master_schedules
            prof_weekly = {}
            if ms is not None and not ms.empty:
                for _, _row in ms.iterrows():
                    _prof = _row.get('교수명')
                    if pd.isna(_prof) or not _prof: continue
                    _n = 0
                    if _row.get('차리생성'): _n += 2  # 차리+판정
                    if _row.get('참관생성'): _n += 1
                    prof_weekly[_prof] = prof_weekly.get(_prof, 0) + _n

            # 보충진료도 카운트 (차리+참관 ≒ 2개)
            for _s in st.session_state.supplementary_schedules:
                _p = _s.get('교수')
                if not _p: continue
                prof_weekly[_p] = prof_weekly.get(_p, 0) + 2

            # 균등 분배 가정 lower bound (임계치 thr=3, 4회+부터 점수):
            # 교수 p에 N_p task, 3K까지는 무료. 초과를 K명에게 균등 분배하면
            # 한 명당 excess ≈ (N_p − 3K)/K → 점수 ≈ (N_p − 3K)² / K
            _wc = st.session_state.get('week_count', 0) or 0
            score_lb = 0.0
            n_over_profs = 0
            for _prof, _w in prof_weekly.items():
                N_p = _w * _wc
                if N_p > 3 * K_gen:
                    excess_total = N_p - 3 * K_gen
                    score_lb += (excess_total ** 2) / K_gen
                    n_over_profs += 1
            score_lb_est = int(math.ceil(score_lb))
            P_count = len(prof_weekly)

            mult = float(st.session_state.cpsat_prof_repeat_multiplier)
            cap_est_lb = int(math.ceil(mult * score_lb_est))
            st.markdown(
                f"📊 **이론적 점수 lower bound** ≈ `{score_lb_est}` "
                f"(균등 분배 가정, 매우 낙관)  ·  "
                f"일반 전공의 {K_gen}명 · 등장 교수 {P_count}명 · 초과 교수 {n_over_profs}명  ·  "
                f"이 값 기준 예상 cap = `{cap_est_lb}`"
            )
            st.caption(
                "⚠️ **이 추정은 슬롯/요일/역할 제약을 모두 무시한 매우 낙관적 하한**입니다. "
                "실제 도달 가능한 `L*`은 보통 더 크며, 솔브 후 결과 화면에 정확히 표시됩니다. "
                "**실제 cap = ⌈배수 × 실제 L*⌉** — 위 추정 cap과 다를 수 있어요."
            )
        except Exception as _e:
            st.caption(f"이론적 추정 불가 (전공의·진료 데이터 부족): `{_e}`")

    # === 로딩 범위 / 부등호(H17) 설정 ===
    # 옛 백업/세션(5 그룹)에서 넘어왔을 때 누락 키 자동 보정 (그룹 4=의국처음 신규)
    _default_lr_full = {0: [4.9, 5.5], 1: [6.3, 6.8], 2: [6.5, 7.2], 3: [7.3, 7.9], 4: [7.5, 8.0], 5: [8.0, 9.0]}
    for _g in range(6):
        if _g not in st.session_state.loading_ranges:
            st.session_state.loading_ranges[_g] = list(_default_lr_full[_g])
    for _g in range(5):
        if _g not in st.session_state.h17_ops:
            st.session_state.h17_ops[_g] = "<"

    with st.expander("⚖️ 로딩 범위 / 그룹 부등호 설정 (고급)", expanded=False):
        st.caption("그룹별 로딩 범위와 인접 그룹 간 부등호를 조정해 CP-SAT를 가동합니다. 기본값은 현재 설정값입니다.")
        if st.button("↩️ 기본값으로 초기화", key="reset_loading_cfg"):
            st.session_state.loading_ranges = {0: [4.9, 5.5], 1: [6.3, 6.8], 2: [6.5, 7.2], 3: [7.3, 7.9], 4: [7.5, 8.0], 5: [8.0, 9.0]}
            st.session_state.h17_ops = {0: "<", 1: "<", 2: "<", 3: "<", 4: "<"}
            for _g in range(6):
                st.session_state.pop(f"lr_lo_{_g}", None)
                st.session_state.pop(f"lr_hi_{_g}", None)
                st.session_state.pop(f"h17_op_{_g}", None)
            st.rerun()
        group_labels = {
            0: "그룹0 · 의국/교육수석 (R3)",
            1: "그룹1 · 학생/진료수석 (R3)",
            2: "그룹2 · 일반 R3",
            3: "그룹3 · R2",
            4: "그룹4 · 의국처음 (R1/R0 + 태그)",
            5: "그룹5 · R1/R0 태그없음",
        }
        op_options = ["<", "<=", "="]
        op_help = "'<' 위 그룹 로딩이 더 큼(엄격) · '<=' 같거나 더 큼 · '=' 두 그룹을 한 그룹으로 병합(로딩 동일 취급)"
        # 폼으로 묶어서, 숫자를 여러 개 고쳐도 [적용]을 누를 때 한 번만 반영된다
        # (0.1씩 올릴 때마다 화면이 새로 그려지는 것 방지)
        with st.form("loading_cfg_form", border=False):
            st.caption("⌨️ 값을 모두 고친 뒤 아래 **[💾 로딩 설정 적용]** 을 눌러야 반영됩니다.")
            _form_ranges = {}
            _form_ops = {}
            for g in range(6):
                lo_cur, hi_cur = st.session_state.loading_ranges[g]
                rc1, rc2, rc3 = st.columns([2, 1, 1])
                rc1.markdown(f"**{group_labels[g]}**")
                lo_new = rc2.number_input("하한", min_value=0.0, max_value=10.0, value=float(lo_cur), step=0.1, key=f"lr_lo_{g}", format="%.1f")
                hi_new = rc3.number_input("상한", min_value=0.0, max_value=10.0, value=float(hi_cur), step=0.1, key=f"lr_hi_{g}", format="%.1f")
                _form_ranges[g] = [lo_new, hi_new]
                if g < 5:
                    op_cur = st.session_state.h17_ops.get(g, "<")
                    idx = op_options.index(op_cur) if op_cur in op_options else 0
                    _form_ops[g] = st.selectbox(
                        f"↕ 그룹{g} 와 그룹{g+1} 관계", options=op_options, index=idx,
                        key=f"h17_op_{g}", help=op_help,
                    )
            _lr_submitted = st.form_submit_button("💾 로딩 설정 적용", type="primary", use_container_width=True)
        if _lr_submitted:
            st.session_state.loading_ranges = _form_ranges
            st.session_state.h17_ops = _form_ops
            st.toast("로딩 범위/부등호가 적용되었습니다.", icon="✅")
            st.rerun()

        # === 장기휴가자 로딩 범위 ===
        st.markdown("---")
        st.markdown("##### 🌴 장기휴가자 로딩 범위")
        st.caption(
            "근무 가능 세션이 **절반 이하**인 사람(휴가가 절반 이상)에게 적용할 로딩 범위입니다. "
            "이들은 연차 내 차이(H16)·그룹 간 부등호(H17)에서 제외되고, 이 범위만 적용됩니다."
        )
        with st.form("longleave_cfg_form", border=False):
            _ll_cur = st.session_state.longleave_range
            lc1, lc2, lc3 = st.columns([2, 1, 1])
            lc1.markdown("**장기휴가자 (가용 ≤ 절반)**")
            _ll_lo = lc2.number_input("하한", min_value=0.0, max_value=10.0,
                                      value=float(_ll_cur[0]), step=0.1, key="ll_lo", format="%.1f")
            _ll_hi = lc3.number_input("상한", min_value=0.0, max_value=10.0,
                                      value=float(_ll_cur[1]), step=0.1, key="ll_hi", format="%.1f")
            _ll_submitted = st.form_submit_button("💾 장기휴가 로딩 적용", type="primary", use_container_width=True)
        if _ll_submitted:
            st.session_state.longleave_range = [_ll_lo, _ll_hi]
            st.toast("장기휴가자 로딩 범위가 적용되었습니다.", icon="✅")
            st.rerun()
        _llc = st.session_state.longleave_range
        st.markdown(f"**현재 적용 중:** `{_llc[0]:.1f} ~ {_llc[1]:.1f}`")
        if _llc[0] > _llc[1]:
            st.warning("하한이 상한보다 큽니다 — 해가 없습니다.")

        # === 보건소/영상 주당 추가 배정 ===
        st.markdown("---")
        st.markdown("##### ➕ 보건소 · 영상 주당 추가 배정")
        st.caption(
            "연건 보건소와 본원 영상은 원래 고정 업무(연보 / 주당 1개)만 받아 로딩이 매우 낮습니다. "
            "필요하면 **1주일에 1개씩** 업무를 더 얹을 수 있습니다."
        )
        st.checkbox(
            "연건 보건소 → **처치(오후)** 를 주당 1개 추가",
            key="weekly_extra_bogeonso",
            help="보건소 담당자에게 매주 처치(오후)를 최소 1개 배정합니다. "
                 "직전휴가 보충(H22)이 있으면 그만큼 더해집니다.",
        )
        st.checkbox(
            "본원 영상 → **예진** 을 주당 1개 추가",
            key="weekly_extra_rad",
            help="영상 파견자에게 매주 예진 1개를 추가로 배정합니다. "
                 "기존 주당 1개(클리닉 또는 처치)는 그대로 유지되어 주당 2개가 됩니다.",
        )

        # === 연차별 판정 개수 범위 (H18) ===
        st.markdown("---")
        st.markdown("##### 📋 연차별 판정 개수 범위 (H18)")
        st.caption(
            "각 전공의가 한 달에 받을 **일반 판정(건증 판정 등)** 개수의 하한~상한입니다. "
            "**R3는 H8 룰로 정확히 1개**라 여기서 조정하지 않습니다(판정=참관 pair와 묶여 있음). "
            "⚠️ `max(R2) < min(R1/R0)` 부등호가 함께 걸리므로, R2 상한을 너무 올리면 해가 없을 수 있습니다."
        )
        _pj_def = {'R2': [3, 20], 'R1': [0, st.session_state.week_count + 5],
                   'R0': [0, st.session_state.week_count + 5], 'week_max': 3}
        for _k, _v in _pj_def.items():
            if _k not in st.session_state.panjung_ranges:
                st.session_state.panjung_ranges[_k] = (list(_v) if isinstance(_v, list) else _v)
        with st.form("panjung_cfg_form", border=False):
            st.caption("⌨️ 값을 모두 고친 뒤 아래 **[💾 판정 범위 적용]** 을 눌러야 반영됩니다.")
            _pj_form = {}
            for _yr, _lbl in [('R2', 'R2'), ('R1', 'R1'), ('R0', 'R0')]:
                _cur = st.session_state.panjung_ranges.get(_yr, _pj_def[_yr])
                pj1, pj2, pj3 = st.columns([2, 1, 1])
                pj1.markdown(f"**{_lbl}**")
                _lo = pj2.number_input("최소", min_value=0, max_value=50, value=int(_cur[0]),
                                       step=1, key=f"pj_lo_{_yr}")
                _hi = pj3.number_input("최대", min_value=0, max_value=50, value=int(_cur[1]),
                                       step=1, key=f"pj_hi_{_yr}")
                _pj_form[_yr] = [_lo, _hi]
            _wk_max = st.number_input(
                "주당 최대 판정 개수 (R2·R1·R0 공통)", min_value=1, max_value=10,
                value=int(st.session_state.panjung_ranges.get('week_max', 3)), step=1, key="pj_week_max",
                help="한 주에 같은 사람이 받을 수 있는 판정 개수 상한",
            )
            _pj_submitted = st.form_submit_button("💾 판정 범위 적용", type="primary", use_container_width=True)
        if _pj_submitted:
            _pj_form['week_max'] = _wk_max
            st.session_state.panjung_ranges = _pj_form
            st.toast("연차별 판정 개수 범위가 적용되었습니다.", icon="✅")
            st.rerun()
        _pjc = st.session_state.panjung_ranges
        st.markdown(
            f"**현재 적용 중:** R3 `1~1(H8 고정)` · "
            f"R2 `{_pjc['R2'][0]}~{_pjc['R2'][1]}` · "
            f"R1 `{_pjc['R1'][0]}~{_pjc['R1'][1]}` · "
            f"R0 `{_pjc['R0'][0]}~{_pjc['R0'][1]}` · 주당 최대 `{_pjc.get('week_max', 3)}`"
        )
        for _yr in ('R2', 'R1', 'R0'):
            if _pjc[_yr][0] > _pjc[_yr][1]:
                st.warning(f"{_yr}: 최소({_pjc[_yr][0]})가 최대({_pjc[_yr][1]})보다 큽니다 — 해가 없습니다.")
        if _pjc['R2'][1] > 0 and _pjc['R1'][1] > 0 and _pjc['R2'][0] >= _pjc['R1'][1]:
            st.warning("R2 최소가 R1 최대 이상입니다 — `max(R2) < min(R1)` 부등호와 충돌해 해가 없을 수 있습니다.")
        # 공급(판정 task 수) 대비 최소 요구량 검사 — 초과하면 무조건 INFEASIBLE
        try:
            _dfa = st.session_state.current_df_all
            _pj_supply = 0
            if _dfa is not None and not _dfa.empty:
                _pj_supply = int(_dfa['task'].apply(
                    lambda t: ('판정' in t) and ('참관' not in t) and ('차리/판정' not in t)).sum())
            _rad_now = {r['이름']: r.get('영상파견요일', []) for r in st.session_state.residents
                        if "본원 영상" in r.get('역할', [])}
            _need = 0
            for _r in st.session_state.residents:
                _ro = _r.get('역할', [])
                if "연건 보건소" in _ro or ("본원 영상" in _ro and _rad_now.get(_r['이름'])):
                    continue                      # 로딩/판정 룰 제외 인원
                if _r['연차'] == 'R3':
                    _need += 1                    # H8: 정확히 1개
                elif _r['연차'] in ('R2', 'R1', 'R0'):
                    _need += int(_pjc[_r['연차']][0])
            if _pj_supply and _need > _pj_supply:
                st.error(
                    f"❌ **최소 요구 {_need}개 > 실제 판정 task {_pj_supply}개** — 이대로는 해가 없습니다(INFEASIBLE). "
                    f"각 연차의 **최소값을 낮추세요.**"
                )
            elif _pj_supply:
                st.caption(f"✅ 최소 요구 {_need}개 / 실제 판정 task {_pj_supply}개 — 공급은 충분합니다.")
        except Exception:
            pass

        st.markdown("---")
        # 현재 적용 중인 값 + 유효성 경고 (폼 밖 — 적용된 값 기준)
        st.markdown("**현재 적용 중인 로딩 설정**")
        _cur_lines = []
        for g in range(6):
            lo_v, hi_v = st.session_state.loading_ranges[g]
            _op = f"  {st.session_state.h17_ops.get(g, '<')}" if g < 5 else ""
            _cur_lines.append(f"- {group_labels[g]}: **{lo_v:.1f} ~ {hi_v:.1f}**{_op}")
            if lo_v > hi_v:
                st.warning(f"{group_labels[g]}: 하한({lo_v})이 상한({hi_v})보다 큽니다 — 해가 없을 수 있습니다.")
        st.markdown("\n".join(_cur_lines))

    if st.button("🚀 자동 스케줄 랜덤 생성 (플랜 B 포함)", use_container_width=True, type="primary"):
        if not st.session_state.residents: st.error("전공의 명단을 등록해주세요.")
        else:
            u_holidays = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
            df_gen = generate_schedule(st.session_state.base_date, st.session_state.week_count, u_holidays, st.session_state.master_schedules, st.session_state.off_slots, supplementary_schedules=st.session_state.supplementary_schedules, biweekly_choice=st.session_state.get('biweekly_choice'))
            rad_days_dict = {
                r['이름']: r.get('영상파견요일', [])
                for r in st.session_state.residents
                if "본원 영상" in r.get('역할', [])
            }

            if st.session_state.use_cpsat:
                # === CP-SAT 모드 ===
                manual_mult = None
                if st.session_state.cpsat_manual_multiplier != "자동":
                    try:
                        manual_mult = float(st.session_state.cpsat_manual_multiplier)
                    except Exception:
                        manual_mult = None
                spinner_msg = f'CP-SAT 솔버 실행 중 (최대 {st.session_state.cpsat_time_limit}초'
                if manual_mult is not None:
                    spinner_msg += f', 배율 {manual_mult:.1f} 고정'
                else:
                    spinner_msg += ', 배율 자동'
                spinner_msg += ')...'
                with st.spinner(spinner_msg):
                    from cpsat_solver import solve_schedule
                    cpsat_result = solve_schedule(
                        df_gen, st.session_state.residents, st.session_state.resident_leaves,
                        st.session_state.week_count, st.session_state.base_date, u_holidays,
                        bogeonso_substitutes=st.session_state.bogeonso_substitutes,
                        rad_days=rad_days_dict,
                        student_practices=st.session_state.student_practices,
                        pain_applicants=st.session_state.pain_applicants,
                        time_limit_sec=st.session_state.cpsat_time_limit,
                        manual_multiplier=manual_mult,
                        loading_ranges=st.session_state.loading_ranges,
                        h17_ops=st.session_state.h17_ops,
                        max_broken_pairs=st.session_state.cpsat_max_broken_pairs,
                        extra_shift_allowance=st.session_state.cpsat_extra_shift,
                        prof_repeat_mode=('eps_constraint' if st.session_state.cpsat_prof_repeat_enabled else 'off'),
                        prof_repeat_multiplier=float(st.session_state.cpsat_prof_repeat_multiplier),
                        prof_repeat_slack=int(st.session_state.cpsat_prof_repeat_slack),
                        shift_forbid_date=st.session_state.get('shift_forbid_date') or None,
                        skip_panjung_obs=st.session_state.get('skip_panjung_obs') or None,
                        assign_prev_month_obs=(not bool(st.session_state.get('relax_prev_month_obs', False))),
                        disabled_rules=st.session_state.get('disabled_rules') or None,
                        manual_pins=_pins_to_task_ids(df_gen, st.session_state.get('manual_pins')),
                        drop_prev_month_obs=bool(st.session_state.get('drop_prev_month_obs', False)),
                        panjung_ranges=st.session_state.get('panjung_ranges') or None,
                        longleave_range=st.session_state.get('longleave_range') or None,
                        weekly_extra_bogeonso=bool(st.session_state.get('weekly_extra_bogeonso', False)),
                        weekly_extra_rad=bool(st.session_state.get('weekly_extra_rad', False)),
                    )

                if cpsat_result['status'] in ['OPTIMAL', 'FEASIBLE']:
                    # 이동 전 원래 날짜 기록 (검증 탭에서 '누락'이 아닌 '이동'으로 인식시키기 위해 필수)
                    _date_ovr = cpsat_result.get('task_date_overrides', {})
                    original_df_dates = {}
                    for tid in _date_ovr:
                        _row = df_gen[df_gen['task_id'] == tid]
                        if not _row.empty:
                            original_df_dates[tid] = _row.iloc[0]['date']
                    st.session_state.cpsat_original_df_dates = original_df_dates
                    # 비고정 task의 시간 결정 적용 (df_gen에 time 컬럼 업데이트)
                    for tid, t_choice in cpsat_result.get('task_time_overrides', {}).items():
                        df_gen.loc[df_gen['task_id'] == tid, 'time'] = t_choice
                    # 날짜 이동 적용 (date_alt로 이동된 차리/판정)
                    weekday_kor = ['월', '화', '수', '목', '금', '토', '일']
                    for tid, new_date in cpsat_result.get('task_date_overrides', {}).items():
                        df_gen.loc[df_gen['task_id'] == tid, 'date'] = new_date
                        # day, week도 업데이트
                        try:
                            new_dt = datetime.strptime(f"{st.session_state.base_date.year}-{new_date}", "%Y-%m-%d").date()
                            new_day = weekday_kor[new_dt.weekday()]
                            new_week = (new_dt - st.session_state.base_date).days // 7 + 1
                            df_gen.loc[df_gen['task_id'] == tid, 'day'] = new_day
                            df_gen.loc[df_gen['task_id'] == tid, 'week'] = new_week
                        except Exception:
                            pass

                    # 전공의별 daily_slots 구성 (공휴일/휴가/메인외래/학생실습/연보/영상 + 배정 결과)
                    # 수동 저장 시에도 동일 함수로 재구성하여 개인별/주차별 뷰가 항상 일치하도록 함
                    res_daily_slots = build_res_daily_slots(
                        cpsat_result['assignments'],
                        cpsat_result.get('forced_assignments', {}),
                        df=df_gen
                    )

                    # 리포트 생성
                    report = ["✅ **CP-SAT 배정 리포트**"]
                    report.append("")
                    phase_used = cpsat_result.get('phase_used', 1)
                    phase_tag = " (1단계: 원래 룰)" if phase_used == 1 else " (2단계: 차리/판정 -1일 이동 허용)"
                    report.append(f"🧮 **솔버 결과**: {cpsat_result['status']} (배율 {cpsat_result.get('multiplier_used'):.2f} 적용{phase_tag}, 소요시간 {cpsat_result['wall_time_sec']:.2f}초)")
                    # 배율/단계 히스토리
                    history = cpsat_result.get('multiplier_history', [])
                    if len(history) > 1:
                        history_strs = []
                        for h in history:
                            tag = f"{h['multiplier']:.2f}-{h.get('phase', 1)}단계({h['status']}, 미배정{h.get('unassigned','?')})"
                            history_strs.append(tag)
                        report.append(f"  - 시도 내역: " + " → ".join(history_strs))
                    report.append("")

                    # === 🎯 S5 교수 중복 결과 (이차 정의: excess² 합) ===
                    _pr_total = cpsat_result.get('prof_repeat_total', None)
                    if _pr_total is not None:
                        _pr_mode = cpsat_result.get('prof_repeat_mode_used', 'soft')
                        _pr_by_res = cpsat_result.get('prof_repeat_by_resident', {}) or {}
                        _pr_pairs = cpsat_result.get('prof_repeat_pairs', []) or []
                        _pr_pair_count = cpsat_result.get('prof_repeat_pair_count', len(_pr_pairs))
                        if _pr_mode.startswith('eps_constraint'):
                            _L = cpsat_result.get('prof_repeat_phase1_L', None)
                            _cap = cpsat_result.get('prof_repeat_cap_used', None)
                            _mult = cpsat_result.get('prof_repeat_multiplier_used', None)
                            _slk = cpsat_result.get('prof_repeat_slack_used', None)
                            _head = (
                                f"🎯 **교수 중복(S5) 강화 모드** — Phase 1 `L* = {_L}`"
                                f", cap = `{_cap}` (= max(⌈×{_mult:.2f}⌉, L+{_slk}))"
                                if _L is not None and _mult is not None
                                else "🎯 **교수 중복(S5) 강화 모드** (ε-constraint)"
                            )
                            if _pr_mode == 'eps_constraint_phase1_only':
                                _head += "  ⚠️ Phase 2 INFEASIBLE → Phase 1 결과로 폴백"
                            elif _pr_mode == 'soft_fallback':
                                _head += "  ⚠️ Phase 1 실패 → 소프트 모드로 폴백"
                            report.append(_head)
                        else:
                            report.append("🎯 **교수 중복(S5)** — 소프트 모드 (가중치 1)")

                        report.append(
                            f"  - **최종 중복 점수** (excess² 합): `{_pr_total}`  ·  임계치 초과 페어 수: `{_pr_pair_count}`"
                        )
                        report.append(
                            "  - 정의: 페어별 점수 = `max(0, count-3)²` → 4회=1·5회=4·6회=9·7회=16"
                        )
                        _nz = {n: v for n, v in _pr_by_res.items() if v > 0}
                        if _nz:
                            report.append("  - 전공의별 중복 점수 (excess² 합):")
                            for _n in sorted(_nz, key=lambda x: (-_nz[x], x)):
                                _np = sum(1 for p in _pr_pairs if p['person'] == _n)
                                report.append(f"    · **{_n}** — 점수 `{_nz[_n]}`, ≥3회 페어 {_np}개")
                            report.append("  - 상세 페어 목록은 아래 **🎯 교수 중복 상세** expander 참고")
                        else:
                            report.append("  - 모든 (전공의, 교수) 페어가 ≤2회 — 중복 없음 ✨")
                        report.append("")
                        # 별도 expander에 DataFrame 출력하기 위해 session_state에 저장
                        st.session_state.prof_repeat_pairs_data = _pr_pairs
                        st.session_state.prof_repeat_by_resident_data = dict(_pr_by_res)
                        st.session_state.prof_repeat_summary = {
                            'total': _pr_total,
                            'pair_count': _pr_pair_count,
                            'mode': _pr_mode,
                            'L_star': cpsat_result.get('prof_repeat_phase1_L'),
                            'cap': cpsat_result.get('prof_repeat_cap_used'),
                            'multiplier': cpsat_result.get('prof_repeat_multiplier_used'),
                            'slack': cpsat_result.get('prof_repeat_slack_used'),
                        }
                    else:
                        # soft 모드 등에서 결과 없음 → 이전 솔브 데이터 제거
                        for _k in ['prof_repeat_pairs_data', 'prof_repeat_by_resident_data', 'prof_repeat_summary']:
                            if _k in st.session_state:
                                del st.session_state[_k]

                    # 사전 검사: 슬롯 부족 날짜 (차리/판정 -1 이동 허용 대상)
                    shortage_dates = cpsat_result.get('shortage_dates', [])
                    if shortage_dates:
                        report.append(f"⚠️ **사전 검사 — 슬롯 부족 날짜 ({len(shortage_dates)}개)** (그 날짜의 차리/판정 -1 평일 이동 허용)")
                        shortage_info = cpsat_result.get('shortage_info', {})
                        for ds in shortage_dates:
                            info = shortage_info.get(ds, {})
                            report.append(f"  - {ds}: task {info.get('tasks','?')}개 / 빈슬롯 {info.get('empty_slots','?')}개")
                        report.append("")

                    # 미배정
                    if cpsat_result['unassigned']:
                        report.append(f"🚨 **미배정 task ({len(cpsat_result['unassigned'])}개)** — 빈 슬롯 부족으로 미배정")
                        task_map = {r['task_id']: r for _, r in df_gen.iterrows()}
                        for tid in cpsat_result['unassigned']:
                            t = task_map.get(tid)
                            if t is not None:
                                report.append(f"  - 주{t['week']} {t['date']}({t['day']}) {t['time']} | {t['task']}")
                        report.append("")

                    # 판정참관 제외 (룰에 따라 '안 하기로 한' 것 — 미배정 아님)
                    _skipped = cpsat_result.get('skipped_panjung_obs', [])
                    if _skipped:
                        report.append(f"🚫 **판정참관 제외 ({len(_skipped)}개)** — 제외 대상자가 건증 판정을 받아 참관을 배정하지 않음 (미배정 아님)")
                        task_map_sk = {r['task_id']: r for _, r in df_gen.iterrows()}
                        for tid in _skipped:
                            t = task_map_sk.get(tid)
                            if t is not None:
                                report.append(f"  - 주{t['week']} {t['date']}({t['day']}) {t['time']} | {t['task']}")
                        report.append("")

                    # 깨진 pairing
                    if cpsat_result['broken_pairs']:
                        report.append(f"🔗 **깨진 pairing ({len(cpsat_result['broken_pairs'])}개)** (최대 5개 허용)")
                        for pid in cpsat_result['broken_pairs']:
                            report.append(f"  - {pid}")
                        report.append("")

                    # 차리/판정 날짜 이동 표시
                    shifted = cpsat_result.get('shifted_tasks', [])
                    if shifted:
                        report.append(f"📅 **날짜 이동된 차리/판정 ({len(shifted)}개)** — 미배정 줄이기 위해 -1 평일 이동")
                        task_map_orig = {r['task_id']: r for _, r in df_gen.iterrows()}
                        date_overrides = cpsat_result.get('task_date_overrides', {})
                        for tid in shifted:
                            t = task_map_orig.get(tid)
                            new_d = date_overrides.get(tid, '?')
                            if t is not None:
                                report.append(f"  - {t['task']}: 원래 → {new_d}로 이동")
                        report.append("")

                    # 사람별 통계
                    from collections import Counter
                    person_count = Counter(cpsat_result['assignments'].values())
                    task_map = {r['task_id']: r for _, r in df_gen.iterrows()}
                    from cpsat_solver import build_problem_data, get_loading_group
                    pdata = build_problem_data(df_gen, st.session_state.residents, st.session_state.resident_leaves,
                                                st.session_state.week_count, st.session_state.base_date, u_holidays,
                                                rad_days=rad_days_dict, student_practices=st.session_state.student_practices,
                                                bogeonso_substitutes=st.session_state.bogeonso_substitutes)
                    forced_per_person = pdata.get('person_forced_count', {})
                    for r in sorted(st.session_state.residents, key=lambda x: get_loading_group(x, rad_days_dict)):
                        name = r['이름']
                        avail = pdata['person_avail_sessions'].get(name, 1)
                        cnt = person_count.get(name, 0)
                        forced = forced_per_person.get(name, 0)  # 강제 연보 슬롯 수
                        total_cnt = cnt + forced  # 일반 task + 연보 (로딩 분자)
                        load = total_cnt * 10 / avail if avail > 0 else 0
                        # 판정/처치/예진 카운트 (H_TX_YEJIN 룰: 처치+예진 합 균등)
                        pj = 0; tx = 0; ye = 0
                        for tid, n in cpsat_result['assignments'].items():
                            if n != name: continue
                            tt = task_map.get(tid)
                            if tt is None: continue
                            if '판정' in tt['task'] and '참관' not in tt['task']:
                                if not (any(p in tt['task'] for p in ['조비룡', '박민선']) and '클리닉' in tt['task']):
                                    pj += 1
                            if '처치' in tt['task']: tx += 1
                            if '예진' in tt['task']: ye += 1
                        report.append(
                            f"- **{name} ({r['연차']})**: 세션 {total_cnt}/{avail} (로딩 {load:.2f})"
                            f" | 🩺판정: **{pj}** | 💉처치+예진: **{tx+ye}** (처치 {tx}/예진 {ye})"
                        )

                    st.session_state.current_df_all = df_gen
                    st.session_state.assignments = cpsat_result['assignments']
                    st.session_state.alloc_report = "\n".join(report)
                    st.session_state.res_daily_slots = res_daily_slots
                    # 수동 저장 시 daily_slots 재구성에 재사용하기 위해 연보(forced) 정보 보관
                    st.session_state.cpsat_forced_assignments = cpsat_result.get('forced_assignments', {})
                    # 제외룰로 드롭된 판정참관 — 검증 탭에서 미배정과 분리해 집계
                    st.session_state.cpsat_skipped_obs = list(cpsat_result.get('skipped_panjung_obs', []))
                    # 이전 달 참관 중 배정 안 된 것 — 별도 집계
                    st.session_state.cpsat_skipped_prev_month_obs = list(cpsat_result.get('skipped_prev_month_obs', []))
                    # 디버깅용 - 백업에 포함시키기 위해 session_state에 저장
                    st.session_state.cpsat_task_time_overrides = cpsat_result.get('task_time_overrides', {})
                    st.session_state.cpsat_task_date_overrides = cpsat_result.get('task_date_overrides', {})
                    st.session_state.cpsat_shifted_tasks = cpsat_result.get('shifted_tasks', [])
                    # FEASIBLE이니까 이전 INFEASIBLE 정보 초기화
                    st.session_state.pop('infeasible_input', None)
                    st.session_state.pop('last_diagnosis', None)
                    st.success(f"CP-SAT 배정 완료! 배율 {cpsat_result.get('multiplier_used'):.2f}, {cpsat_result['wall_time_sec']:.2f}초")
                    st.rerun()
                else:
                    # INFEASIBLE — 진단은 별도 버튼으로 사용자 명시적 실행
                    error_lines = ["❌ **CP-SAT 솔버: 해가 없습니다 (INFEASIBLE)**"]
                    error_lines.append("")
                    error_lines.append(f"배율 {cpsat_result.get('multiplier_used', '?')}로 모든 hard rule을 동시 만족하는 배정이 존재하지 않습니다.")
                    error_lines.append("")
                    error_lines.append("**가능한 해결책:**")
                    error_lines.append("- 전공의 인원 추가")
                    error_lines.append("- 휴가 조정")
                    error_lines.append("- 아래 **'INFEASIBLE 진단 실행'** 버튼으로 충돌 룰 찾기 (시간 걸림)")
                    error_lines.append("- 위 충돌 룰 완화 (개발자 문의)")
                    error_lines.append("")
                    history = cpsat_result.get('multiplier_history', [])
                    if history:
                        error_lines.append("**시도 내역:**")
                        for h in history:
                            error_lines.append(f"  - 배율 {h['multiplier']:.2f}: {h['status']} ({h['wall_time']:.2f}초)")
                    st.error("\n".join(error_lines))
                    st.session_state.alloc_report = "\n".join(error_lines)
                    # 진단 버튼을 위해 입력 데이터 보관
                    st.session_state.infeasible_input = {
                        'df_gen': df_gen,
                        'residents': st.session_state.residents,
                        'resident_leaves': st.session_state.resident_leaves,
                        'week_count': st.session_state.week_count,
                        'base_date': st.session_state.base_date,
                        'holidays': u_holidays,
                        'bogeonso_substitutes': st.session_state.bogeonso_substitutes,
                        'rad_days': rad_days_dict,
                        'student_practices': st.session_state.student_practices,
                        'pain_applicants': st.session_state.pain_applicants,
                        'auto_mult': cpsat_result.get('multiplier_used', 1.0),
                        'loading_ranges': st.session_state.loading_ranges,
                        'h17_ops': st.session_state.h17_ops,
                        'max_broken_pairs': st.session_state.cpsat_max_broken_pairs,
                    }

            else:
                # === 기존 점수 방식 (호환성) ===
                with st.spinner('배정 중... (10회 시도 후 최적 선택)'):
                    mult = diagnosis_result['multiplier'] if diagnosis_result else 1.0
                    from utils import run_auto_assignment_multi
                    o_df, n_assign, r_txt, n_slots = run_auto_assignment_multi(
                        df_gen, st.session_state.residents, st.session_state.resident_leaves,
                        st.session_state.week_count, st.session_state.base_date, u_holidays,
                        st.session_state.pain_applicants, st.session_state.student_practices,
                        bogeonso_substitutes=st.session_state.bogeonso_substitutes,
                        rad_days=rad_days_dict,
                        target_mult_multiplier=mult,
                        num_trials=10
                    )
                    st.session_state.current_df_all, st.session_state.assignments, st.session_state.alloc_report, st.session_state.res_daily_slots = o_df, n_assign, r_txt, n_slots
                st.success("자동 배정 완료! (10회 시도 중 최적 결과 선택)")
                st.rerun()
    # === 옛 백업 회복: alloc_report 비어있고 assignments는 있으면 처치/판정/페어/S5 재계산 ===
    if (
        not st.session_state.alloc_report
        and st.session_state.assignments
        and not st.session_state.current_df_all.empty
        and not st.session_state.get('prof_repeat_summary')
    ):
        st.warning(
            "📋 배정 리포트가 비어있습니다 (옛 백업이거나 요약이 저장되지 않은 상태). "
            "현재 배정에서 **처치/판정/예진 카운트, 깨진 pairing, 미배정 task, S5 중복지수**를 재계산할 수 있습니다. "
            "Phase 1 `L*`/cap만 다시 솔브해야 알 수 있어요."
        )
        if st.button("🔄 현재 배정에서 요약 전체 재계산", key="recompute_all"):
            from cpsat_solver import _extract_prof_category
            _df_now = st.session_state.current_df_all
            _task_map = {r['task_id']: r['task'] for _, r in _df_now.iterrows()}
            _assigns = st.session_state.assignments

            # 1) 전공의별 판정/처치/예진/total 카운트
            _person_stats = {r['이름']: {'pj': 0, 'tx': 0, 'ye': 0, 'total': 0} for r in st.session_state.residents}
            for _tid, _person in _assigns.items():
                _tn = _task_map.get(_tid, '')
                if _person not in _person_stats:
                    _person_stats[_person] = {'pj': 0, 'tx': 0, 'ye': 0, 'total': 0}
                _person_stats[_person]['total'] += 1
                if '판정' in _tn and '참관' not in _tn:
                    if not (any(p in _tn for p in ['조비룡', '박민선']) and '클리닉' in _tn):
                        _person_stats[_person]['pj'] += 1
                if '처치' in _tn:
                    _person_stats[_person]['tx'] += 1
                if '예진' in _tn:
                    _person_stats[_person]['ye'] += 1

            # 2) 깨진 pairing 재계산 (같은 pair_id가 두 사람 이상에 배정되면 broken)
            _pair_groups = {}
            for _, _row in _df_now.iterrows():
                _pid = _row.get('pair_id', '')
                if not _pid or (isinstance(_pid, float) and pd.isna(_pid)):
                    continue
                _pair_groups.setdefault(_pid, []).append((_row['task_id'], _assigns.get(_row['task_id'])))
            _broken = []
            for _pid, _items in _pair_groups.items():
                _persons = {p for _, p in _items if p is not None}
                if len(_persons) > 1:
                    _broken.append(_pid)

            # 3) 미배정 task (판정참관 제외룰로 '안 하기로 한' 것은 미배정에서 분리)
            _skip_ids = set(st.session_state.get('cpsat_skipped_obs', []))
            _unassigned_ids = [r['task_id'] for _, r in _df_now.iterrows()
                               if r['task_id'] not in _assigns and r['task_id'] not in _skip_ids]
            _skipped_ids_now = [r['task_id'] for _, r in _df_now.iterrows()
                                if r['task_id'] not in _assigns and r['task_id'] in _skip_ids]

            # 4) S5 페어 (이차 정의, 일반 임계치 thr=3)
            _person_pair = {}
            for _tid, _person in _assigns.items():
                _tn = _task_map.get(_tid, '')
                _prof, _cat = _extract_prof_category(_tn)
                if _prof is None:
                    continue
                _key = (_person, _prof, _cat)
                _person_pair[_key] = _person_pair.get(_key, 0) + 1
            _pairs = []; _by_res = {}; _total = 0
            for (_p, _prof, _cat), _cnt in _person_pair.items():
                _thr = 3
                if _cnt > _thr:
                    _ex = _cnt - _thr; _sc = _ex * _ex
                    _pairs.append({'person': _p, 'prof': _prof, 'category': _cat,
                                   'count': _cnt, 'excess': _ex, 'score': _sc, 'threshold': _thr})
                    _by_res[_p] = _by_res.get(_p, 0) + _sc
                    _total += _sc

            # 5) alloc_report 텍스트 재구성
            _report = ["📋 **재계산된 배정 요약** (백업에서 복원, Phase 1 `L*`/cap은 다시 솔브 필요)", ""]
            _skip_note = f" / 판정참관 제외 {len(_skipped_ids_now)}건" if _skipped_ids_now else ""
            _report.append(f"📊 **배정**: 총 {len(_assigns)}건 / 미배정 {len(_unassigned_ids)}건{_skip_note} / 깨진 pairing {len(_broken)}개")
            _report.append("")
            _report.append("**전공의별 카운트**")
            for _r in st.session_state.residents:
                _n = _r['이름']
                _s = _person_stats.get(_n)
                if not _s or _s['total'] == 0:
                    continue
                _report.append(
                    f"- **{_n} ({_r['연차']})**: 세션 {_s['total']}"
                    f" | 🩺판정 **{_s['pj']}** | 💉처치+예진 **{_s['tx']+_s['ye']}** (처치 {_s['tx']}/예진 {_s['ye']})"
                )
            _report.append("")
            if _broken:
                _report.append(f"🔗 **깨진 pairing ({len(_broken)}개)**")
                for _pid in _broken[:50]:
                    _report.append(f"  - {_pid}")
                if len(_broken) > 50:
                    _report.append(f"  - ... 외 {len(_broken) - 50}개")
                _report.append("")
            if _unassigned_ids:
                _report.append(f"🚨 **미배정 task ({len(_unassigned_ids)}개)**")
                for _tid in _unassigned_ids[:50]:
                    _report.append(f"  - {_task_map.get(_tid, _tid)}")
                if len(_unassigned_ids) > 50:
                    _report.append(f"  - ... 외 {len(_unassigned_ids) - 50}개")
                _report.append("")
            if _skipped_ids_now:
                _report.append(f"🚫 **판정참관 제외 ({len(_skipped_ids_now)}개)** — 룰에 따라 배정하지 않음 (미배정 아님)")
                for _tid in _skipped_ids_now[:50]:
                    _report.append(f"  - {_task_map.get(_tid, _tid)}")
                if len(_skipped_ids_now) > 50:
                    _report.append(f"  - ... 외 {len(_skipped_ids_now) - 50}개")
                _report.append("")
            st.session_state.alloc_report = "\n".join(_report)

            # 6) S5 expander용 session 채움
            st.session_state.prof_repeat_pairs_data = _pairs
            st.session_state.prof_repeat_by_resident_data = _by_res
            st.session_state.prof_repeat_summary = {
                'total': _total, 'pair_count': len(_pairs),
                'mode': 'recomputed', 'L_star': None, 'cap': None,
                'multiplier': None, 'slack': None,
            }
            st.toast(f"✅ 요약 재계산 — 깨진 pair {len(_broken)}, 미배정 {len(_unassigned_ids)}, S5 페어 {len(_pairs)}", icon="✅")
            st.rerun()

    if st.session_state.alloc_report:
        with st.expander("📊 배정 리포트", expanded=True): st.info(st.session_state.alloc_report)

    # === 🎯 교수 중복(S5) 상세 — 페어 DataFrame ===
    if st.session_state.get('prof_repeat_pairs_data') is not None:
        _pairs = st.session_state.get('prof_repeat_pairs_data', [])
        _byres = st.session_state.get('prof_repeat_by_resident_data', {})
        _sum = st.session_state.get('prof_repeat_summary', {})
        with st.expander("🎯 교수 중복 (S5) 상세", expanded=True):
            # 요약
            sc1, sc2, sc3, sc4 = st.columns(4)
            sc1.metric("총 점수 (excess² 합)", _sum.get('total', 0))
            sc2.metric("≥3회 페어 수", _sum.get('pair_count', 0))
            if _sum.get('L_star') is not None:
                sc3.metric("Phase 1 L*", _sum.get('L_star'))
                _cap_v = _sum.get('cap')
                _mult_v = _sum.get('multiplier')
                _slk_v = _sum.get('slack')
                if _mult_v is not None and _slk_v is not None:
                    sc4.metric("cap", f"{_cap_v}", help=f"= max(⌈×{_mult_v:.2f}⌉, L+{_slk_v})")
                else:
                    sc4.metric("cap", _cap_v)
            else:
                sc3.metric("모드", _sum.get('mode', 'soft'))
            st.caption(
                    "정의: 페어별 점수 = `max(0, count-3)²` → **4회=1 · 5회=4 · 6회=9 · 7회=16** "
                    "(4회+부터 카운트, 분산 강한 페널티)"
                )

            # 전공의별
            _nz = {n: v for n, v in _byres.items() if v > 0}
            if _nz:
                st.markdown("##### 👥 전공의별 중복 점수")
                _res_rows = []
                for _n in sorted(_nz, key=lambda x: (-_nz[x], x)):
                    _np = sum(1 for p in _pairs if p['person'] == _n)
                    _res_rows.append({"전공의": _n, "중복 점수": _nz[_n], "≥3회 페어 수": _np})
                st.dataframe(pd.DataFrame(_res_rows), use_container_width=True, hide_index=True)

            # 페어 상세
            if _pairs:
                st.markdown("##### 🔗 페어 상세 (회수 내림차순)")
                st.caption(
                    "카테고리: **외래**(외래·암외래 통합) / **건증**(판정·판정참관) / **통증클리닉** 등 — "
                    "같은 교수라도 카테고리가 다르면 별개 페어로 카운트합니다."
                )
                _pair_rows = []
                for _p in sorted(_pairs, key=lambda x: (-x['score'], -x['count'], x['person'])):
                    _pair_rows.append({
                        "전공의": _p['person'],
                        "교수": f"Pf. {_p['prof']}",
                        "카테고리": _p.get('category', '?'),
                        "임계치": _p.get('threshold', 3),
                        "회수": _p['count'],
                        "excess (count-thr)": _p['excess'],
                        "점수 (excess²)": _p['score'],
                    })
                _df_pairs = pd.DataFrame(_pair_rows)

                def _color_count(v):
                    try: iv = int(v)
                    except Exception: return ''
                    if iv >= 6: return 'background-color: #c0392b; color: white; font-weight: bold;'
                    if iv == 5: return 'background-color: #e67e22; color: white; font-weight: bold;'
                    if iv == 4: return 'background-color: #f39c12; color: black; font-weight: bold;'
                    if iv == 3: return 'background-color: #f7dc6f; color: black;'
                    return ''
                try:
                    _styled = _df_pairs.style.applymap(_color_count, subset=['회수'])
                    st.dataframe(_styled, use_container_width=True, hide_index=True)
                except Exception:
                    st.dataframe(_df_pairs, use_container_width=True, hide_index=True)
            else:
                st.success("✨ 중복 없음 — 모든 (전공의, 교수) 페어가 ≤2회")

    # === INFEASIBLE 진단 버튼 ===
    if st.session_state.get('infeasible_input'):
        with st.expander("🔍 INFEASIBLE 진단", expanded=True):
            st.warning("배정 실패 (INFEASIBLE) — 어떤 룰이 충돌하는지 진단할 수 있습니다.")
            diag_col1, diag_col2, diag_col3 = st.columns([1, 1, 2])
            with diag_col1:
                diag_per_solve = st.number_input("각 시도 제한 (초)", min_value=3, max_value=30, value=5, step=1,
                                                  help="각 룰 조합 시도당 시간. 짧을수록 빠르지만 정확도 낮음.")
            with diag_col2:
                enable_drop_two = st.checkbox("2개 빼기 단계", value=False,
                                               help="단일 룰/1개 빼기로 원인 못 찾으면 2개씩 빼서 시도 (~30회 추가, 시간 많이 소요)")
            with diag_col3:
                est_time = 11 * diag_per_solve + 1 * diag_per_solve + 11 * diag_per_solve
                if enable_drop_two:
                    est_time += 55 * diag_per_solve
                st.caption(f"예상 시간: ~{est_time}초")
            if st.button("🔍 진단 실행", use_container_width=True, type="primary"):
                ii = st.session_state.infeasible_input
                with st.spinner(f"진단 중... 최대 {est_time}초"):
                    from cpsat_solver import diagnose_infeasibility
                    try:
                        diag = diagnose_infeasibility(
                            ii['df_gen'], ii['residents'], ii['resident_leaves'],
                            ii['week_count'], ii['base_date'], ii['holidays'],
                            bogeonso_substitutes=ii['bogeonso_substitutes'],
                            rad_days=ii['rad_days'],
                            student_practices=ii['student_practices'],
                            pain_applicants=ii['pain_applicants'],
                            target_mult_multiplier=ii['auto_mult'],
                            per_solve_time=diag_per_solve,
                            enable_drop_two=enable_drop_two,
                            loading_ranges=ii.get('loading_ranges'),
                            h17_ops=ii.get('h17_ops'),
                            max_broken_pairs=ii.get('max_broken_pairs', 5),
                        )
                        st.session_state.last_diagnosis = diag
                    except Exception as e:
                        st.error(f"진단 실패: {e}")
                        st.session_state.last_diagnosis = None
                st.rerun()

            # 진단 결과 표시
            diag = st.session_state.get('last_diagnosis')
            if diag:
                st.markdown("---")
                st.markdown("### 🔍 진단 결과")
                method = diag.get('method', 'unknown')
                note = diag.get('note', '')
                if note:
                    st.info(note)
                blocking = diag.get('blocking_rules', [])
                descs = diag.get('rule_descriptions', {})
                if blocking == ['BASE']:
                    st.error("**기본 슬롯/사람 제약 자체로 INFEASIBLE** — 인원 부족 또는 휴가 과다")
                elif blocking:
                    st.markdown(f"**충돌 룰 ({len(blocking)}개):**")
                    for r in blocking:
                        st.markdown(f"- {descs.get(r, r)}")
                    if method == 'single_rule_block':
                        st.info("→ 각 룰이 단독으로 INFEASIBLE을 유발. 위 룰 중 하나 완화 필요.")
                    elif method == 'drop_one':
                        st.info("→ 위 룰 중 **1개만 빼도 풀이 가능**. 우선순위 낮은 룰 양보 권장.")
                    elif method == 'drop_one_no_result':
                        st.warning("→ 1개씩 빼기로는 원인 못 찾음. **'2개 빼기 단계'** 체크하고 재시도.")
                    elif method == 'drop_two':
                        pair_combos = diag.get('pair_combinations', [])
                        st.markdown("→ **2개 룰을 동시에 빼야 풀이 가능**. 예시 조합:")
                        for p in pair_combos[:3]:
                            st.markdown(f"  · {p[0]} + {p[1]}")
    assign_df = st.session_state.current_df_all.copy()
    assign_df['배정된_전공의'] = assign_df['task_id'].map(lambda x: st.session_state.assignments.get(x, ""))
    res_choices = [""] + [r["이름"] for r in st.session_state.residents]
    edited_assign = st.data_editor(assign_df[['week', 'date', 'day', 'time', 'prof', 'task', '배정된_전공의', 'task_id']], column_config={"task_id": None, "배정된_전공의": st.column_config.SelectboxColumn("배정", options=res_choices)}, use_container_width=True, hide_index=True, height=500)
    if st.button("💾 수동 배정 저장", use_container_width=True):
        st.session_state.assignments = {row['task_id']: row['배정된_전공의'] for _, row in edited_assign.iterrows() if row['배정된_전공의']}
        # 개인별/주차별 현황 뷰가 읽는 res_daily_slots를 새 배정으로 재구성 (불일치 방지)
        st.session_state.res_daily_slots = build_res_daily_slots(
            st.session_state.assignments,
            st.session_state.get('cpsat_forced_assignments', {})
        )
        st.success("저장 완료!")
        st.rerun()

# 탭 인덱스 7 (기존 6): 전공의 개인별
with tabs[7]:
    sorted_res_list = get_sorted_residents(st.session_state.residents)
    task_map = st.session_state.current_df_all.set_index('task_id')['task'].to_dict()
    for i in range(0, len(sorted_res_list), 2):
        cols = st.columns(2)
        for j in range(2):
            if i + j < len(sorted_res_list):
                res = sorted_res_list[i + j]; res_name = res['이름']
                with cols[j]:
                    st.markdown(f"#### 👤 {res['연차']} {res_name}")
                    html = "<table style='width:100%; font-size:0.7rem; text-align:center; border-collapse:collapse; table-layout:fixed;'><tr style='background-color:#444; color:white;'><th>월</th><th>화</th><th>수</th><th>목</th><th>금</th></tr>"
                    for w in range(1, st.session_state.week_count + 1):
                        for time in ["오전", "오후"]:
                            html += "<tr>"
                            for d_idx, day in enumerate(["월", "화", "수", "목", "금"]):
                                d_str = (st.session_state.base_date + timedelta(days=(w-1)*7 + d_idx)).strftime("%m-%d")
                                bg, fg, content = "white", "black", ""
                                if d_str in user_holidays: bg, content = "#D3D3D3", "공휴일"
                                else:
                                    a_id = st.session_state.res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get(time)
                                    if a_id:
                                        if any(kw in a_id for kw in ["휴가", "Off"]): bg, fg, content = "#A6A6A6", "white", a_id
                                        elif a_id == "영상": bg, content = "#D5E8D4", a_id
                                        elif any(kw in a_id for kw in ["메인외래", "연보"]): bg, content = "#E8F8F5" if "메인" in a_id else "#FF85FF", a_id
                                        else: content = task_map.get(a_id, a_id); bg, fg = get_task_style(content)
                                    else:
                                        at = [row['task'] for tid, name in st.session_state.assignments.items() if name == res_name for _, row in df_all[df_all['task_id'] == tid].iterrows() if row['date'] == d_str and row['time'] == time]
                                        if at: content = at[0]; bg, fg = get_task_style(content)
                                html += f"<td style='border:1px solid #ddd; padding:3px; vertical-align:top;'><div style='font-size:0.55rem; color:#888; text-align:left;'>{d_str if time=='오전' else '&nbsp;'}</div><div style='background-color:{bg}; color:{fg}; padding:2px; border-radius:3px; min-height:35px; display:flex; align-items:center; justify-content:center; font-weight:bold; line-height:1.1;'>{content}</div></td>"
                            html += "</tr>"
                    st.markdown(html + "</table><br>", unsafe_allow_html=True)

# 탭 인덱스 8 (기존 7): 주차별 전체 현황
with tabs[8]:
    st.subheader("🗓️ 주차별 전공의 전체 스케줄 (Excel 출력 양식)")
    sorted_res_list = get_sorted_residents(st.session_state.residents)
    task_map = st.session_state.current_df_all.set_index('task_id')['task'].to_dict()
    if not sorted_res_list: st.warning("등록된 전공의가 없습니다.")
    else:
        try:
            _dl_col1, _dl_col2 = st.columns([2, 3])
            with _dl_col2:
                st.checkbox(
                    "빈 5주차 포함 (양식 통일용)",
                    key='excel_pad_week5_ck_1',
                    on_change=_sync_pad5, args=('excel_pad_week5_ck_1',),
                    help="스케줄이 4주여도 다운로드 엑셀에 빈 5주차 행을 padding합니다. 스케줄 자체엔 영향 없음 (엑셀에만 적용)."
                )
            _res_sig = _excel_sig(
                'res', st.session_state.week_count, st.session_state.base_date,
                bool(st.session_state.excel_pad_week5), user_holidays,
                st.session_state.assignments, st.session_state.res_daily_slots,
                [r['이름'] for r in sorted_res_list],
            )
            _excel_download_ui(
                _dl_col1, '_xls_res_sig', '_xls_res_data', _res_sig,
                lambda: generate_excel_data(
                    st.session_state.week_count, st.session_state.base_date, sorted_res_list,
                    user_holidays, st.session_state.res_daily_slots, st.session_state.assignments,
                    st.session_state.current_df_all, task_map,
                    include_blank_week5=bool(st.session_state.excel_pad_week5)),
                "📥 이 양식 그대로 엑셀 다운로드",
                f"의국_스케줄_{datetime.today().strftime('%Y%m%d')}.xlsx",
                'mk_res_xls',
            )
        except Exception as e:
            st.error(f"엑셀 생성 중 오류가 발생했습니다: {str(e)}"); st.code(traceback.format_exc())
    for w in range(1, st.session_state.week_count + 1):
        st.markdown(f"#### 📅 {w}주차 현황")
        html = f"<table style='width:100%; font-size:0.8rem; text-align:center; border-collapse:collapse; border: 2px solid #333; table-layout:fixed; font-family:\"맑은 고딕\";'>"
        html += "<tr style='background-color:#F2F2F2;'> <th style='border:1px solid #333; width:6%;'>주차</th> <th style='border:1px solid #333; width:10%;'>성명/역할</th>"
        for d_idx, day_name in enumerate(["월", "화", "수", "목", "금"]):
            dt = st.session_state.base_date + timedelta(days=(w-1)*7 + d_idx)
            html += f"<th style='border:1px solid #333;'>{dt.strftime('%m월 %d일')}({day_name})</th>"
        html += "</tr>"
        prev_group = None
        for res in sorted_res_list:
            res_name = res['이름']; roles = ", ".join(res['역할']) if res['역할'] else "&nbsp;"
            curr_group = "R1/R0" if res['연차'] in ["R1", "R0"] else res['연차']
            is_new_group = prev_group is not None and curr_group != prev_group
            prev_group = curr_group
            top_border = "2px solid #000" if is_new_group else "1px solid #333"
            html += f"<tr>"
            if res == sorted_res_list[0]: html += f"<td rowspan='{len(sorted_res_list)*2}' style='border:1px solid #333; background-color:#fff;'>{w}주차</td>"
            html += f"<td style='border-top:{top_border}; border-left:1px solid #333; border-right:1px solid #333; border-bottom:1px solid #333; font-weight:bold; background-color:#fff;'>{res_name}</td>"
            for d_idx in range(5):
                d_str = (st.session_state.base_date + timedelta(days=(w-1)*7 + d_idx)).strftime("%m-%d")
                am_c, pm_c, am_bg, pm_bg, am_fg, pm_fg = "", "", "white", "white", "black", "black"
                if d_str in user_holidays: am_c = pm_c = "공휴일"; am_bg = pm_bg = "#A6A6A6"
                else:
                    aid = st.session_state.res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get('오전')
                    if aid:
                        if any(kw in aid for kw in ["휴가", "Off"]): am_c, am_bg, am_fg = aid, "#A6A6A6", "white"
                        elif aid == "영상": am_c, am_bg = aid, "#D5E8D4"
                        elif any(kw in aid for kw in ["메인외래", "연보"]): am_c, am_bg = aid, ("#E8F8F5" if "메인" in aid else "#FF85FF")
                        else: am_c = task_map.get(aid, aid); am_bg, am_fg = get_task_style(am_c)
                    pid = st.session_state.res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get('오후')
                    if pid:
                        if any(kw in pid for kw in ["휴가", "Off"]): pm_c, pm_bg, pm_fg = pid, "#A6A6A6", "white"
                        elif pid == "영상": pm_c, pm_bg = pid, "#D5E8D4"
                        elif any(kw in pid for kw in ["메인외래", "연보"]): pm_c, pm_bg = pid, ("#E8F8F5" if "메인" in pid else "#FF85FF")
                        else: pm_c = task_map.get(pid, pid); pm_bg, pm_fg = get_task_style(pm_c)
                html += f"<td style='border-top:{top_border}; border-left:1px solid #333; border-right:1px solid #333; border-bottom:1px solid #333; background-color:{am_bg}; color:{am_fg}; font-weight:bold;'>{am_c}</td>"
            html += "</tr>"
            html += f"<tr><td style='border:1px solid #333; font-size:0.7rem; background-color:#fff;'>{roles}</td>"
            for d_idx in range(5):
                d_str = (st.session_state.base_date + timedelta(days=(w-1)*7 + d_idx)).strftime("%m-%d")
                am_c, pm_c, pm_bg, pm_fg = "", "", "white", "black"
                if d_str in user_holidays: am_c = pm_c = "공휴일"; pm_bg = "#A6A6A6"
                else:
                    aid = st.session_state.res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get('오전')
                    pid = st.session_state.res_daily_slots.get(res_name, {}).get('daily_slots', {}).get(d_str, {}).get('오후')
                    if aid: am_c = task_map.get(aid, aid) if aid not in ["휴가","Off","메인외래","연보","영상"] else aid
                    if pid:
                        if any(kw in pid for kw in ["휴가", "Off"]): pm_c, pm_bg, pm_fg = pid, "#A6A6A6", "white"
                        elif pid == "영상": pm_c, pm_bg = pid, "#D5E8D4"
                        elif any(kw in pid for kw in ["메인외래", "연보"]): pm_c, pm_bg = pid, ("#E8F8F5" if "메인" in pid else "#FF85FF")
                        else: pm_c = task_map.get(pid, pid); pm_bg, pm_fg = get_task_style(pm_c)
                html += f"<td style='border:1px solid #333; background-color:{pm_bg}; color:{pm_fg}; font-weight:bold;'>{pm_c}</td>"
            html += "</tr>"
        html += "</table><br>"
        st.write(html, unsafe_allow_html=True)

# 탭 인덱스 1 (기존 0): 교수별 시간표
with tabs[1]:
    # === 엑셀 다운로드 ===
    try:
        _pdl_col1, _pdl_col2 = st.columns([2, 3])
        with _pdl_col2:
            st.checkbox(
                "빈 5주차 포함 (양식 통일용)",
                key='excel_pad_week5_ck_2',
                on_change=_sync_pad5, args=('excel_pad_week5_ck_2',),
                help="스케줄이 4주여도 다운로드 엑셀에 빈 5주차 행을 padding합니다. 스케줄 자체엔 영향 없음 (엑셀에만 적용)."
            )
        _prof_sig = _excel_sig(
            'prof', st.session_state.week_count, st.session_state.base_date,
            bool(st.session_state.excel_pad_week5), user_holidays,
            st.session_state.off_slots, st.session_state.supplementary_schedules,
            st.session_state.master_schedules.to_dict(orient='records'),
        )
        _excel_download_ui(
            _pdl_col1, '_xls_prof_sig', '_xls_prof_data', _prof_sig,
            lambda: generate_prof_schedule_excel(
                st.session_state.week_count,
                st.session_state.base_date,
                st.session_state.master_schedules,
                st.session_state.supplementary_schedules,
                user_holidays,
                st.session_state.off_slots,
                PROF_ORDER,
                include_blank_week5=bool(st.session_state.excel_pad_week5),
                biweekly_choice=st.session_state.get('biweekly_choice')),
            "📥 교수별 시간표 엑셀 다운로드",
            f"교수별_시간표_{datetime.today().strftime('%Y%m%d')}.xlsx",
            'mk_prof_xls',
        )
    except Exception as _e:
        st.caption(f"엑셀 생성 불가: `{_e}`")

    for i in range(0, len(PROF_ORDER), 2):
        cols = st.columns(2)
        for j in range(2):
            if i + j < len(PROF_ORDER):
                p_name = PROF_ORDER[i+j]
                with cols[j]:
                    st.markdown(f"#### 🏥 Pf. {p_name}")
                    h = "<table style='width:100%; font-size:0.75rem; text-align:center; border-collapse:collapse; table-layout:fixed;'><tr style='background-color:#333; color:white;'><th>월</th><th>화</th><th>수</th><th>목</th><th>금</th></tr>"
                    for w in range(1, st.session_state.week_count + 1):
                        for time in ["오전", "오후"]:
                            h += "<tr>"
                            for d_idx, day in enumerate(["월", "화", "수", "목", "금"]):
                                d_str = (st.session_state.base_date + timedelta(days=(w-1)*7 + d_idx)).strftime("%m-%d")
                                is_h, is_o = d_str in user_holidays, is_prof_off(st.session_state.off_slots, p_name, d_str, time)
                                ci = ""
                                has_chari = True; has_chamgwan = True
                                for _, r in st.session_state.master_schedules.iterrows():
                                    if pd.isna(r["교수명"]): continue
                                    if r["교수명"] == p_name and r["요일"] == day and r["시간"] == time:
                                        _clv = str(r["진료명"]) if pd.notna(r["진료명"]) else ""
                                        if biweekly_week_active(str(r["주기"]), w, f"{p_name}|{day}|{time}|{_clv}", st.session_state.get("biweekly_choice")):
                                            ci = r["진료명"]
                                            has_chari = bool(r.get("차리생성", True))
                                            has_chamgwan = bool(r.get("참관생성", True))
                                # 보충진료 확인 (보통 차리+참관 모두 있음)
                                if not ci:
                                    for s in st.session_state.supplementary_schedules:
                                        if s["교수"] == p_name and s["날짜"] == d_str and s["시간"] == time:
                                            ci = s["진료명"]
                                            has_chari = True; has_chamgwan = True
                                            break
                                bg, fg, bd = get_prof_raw_style(ci, is_o, is_h, False, prof=p_name, has_chari=has_chari, has_chamgwan=has_chamgwan)
                                h += f"<td style='border:1px solid #ddd; padding:4px; vertical-align:top;'><div style='height:14px; font-size:0.6rem; color:#888; text-align:left;'>{d_str if time=='오전' else '&nbsp;'}</div><div style='background-color:{bg}; color:{fg}; padding:2px; border-radius:3px; min-height:38px; display:flex; align-items:center; justify-content:center; border:{bd}; font-weight:bold;'>{is_h and '공휴일' or (is_o and '휴진' or ci)}</div></td>"
                            h += "</tr>"
                    st.write(h + "</table><br>", unsafe_allow_html=True)

# 탭 인덱스 2 (기존 1): 주차별 가이드
with tabs[2]:
    for w in range(1, st.session_state.week_count + 1):
        st.subheader(f"📍 {w}주차")
        w_df = st.session_state.current_df_all[st.session_state.current_df_all['week'] == w]; cols = st.columns(5)
        for i, day in enumerate(["월", "화", "수", "목", "금"]):
            with cols[i]:
                st.markdown(f"<div style='text-align:center; font-weight:bold; border-bottom:2px solid black;'>{day}요일</div>", unsafe_allow_html=True)
                for t in ["오전", "오후"]:
                    st.markdown(f"<div style='background-color:#F2F2F2; padding:2px; font-size:0.7rem; font-weight:bold; text-align:center; margin-top:5px;'>{t}</div>", unsafe_allow_html=True)
                    tasks = w_df[(w_df['day'] == day) & (w_df['time'] == t)]
                    for _, row in tasks.iterrows():
                        bg, fg = get_task_style(row['task']); assignee = st.session_state.assignments.get(row['task_id'], "")
                        st.markdown(f"<div style='background-color:{bg}; color:{fg}; padding:3px; border-radius:3px; margin-bottom:2px; border:1px solid #ddd; font-size:0.7rem;'>{row['task']} <b>({assignee})</b></div>" if assignee else f"<div style='background-color:{bg}; color:{fg}; padding:3px; border-radius:3px; margin-bottom:2px; border:1px solid #ddd; font-size:0.7rem;'>{row['task']}</div>", unsafe_allow_html=True)

# 탭 인덱스 3 (기존 2): 규칙 설정 (수정된 st.form 적용 완료)
with tabs[3]:
    with st.form("master_schedule_form"):
        edited_df = st.data_editor(
            st.session_state.master_schedules,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic"
        )
        submitted = st.form_submit_button("🔄 규칙 적용")
        if submitted:
            st.session_state.master_schedules = edited_df
            user_holidays = [h.strip() for h in st.session_state.user_holidays_str.split(',') if h.strip()]
            st.session_state.current_df_all = generate_schedule(
                st.session_state.base_date,
                st.session_state.week_count,
                user_holidays,
                st.session_state.master_schedules,
                st.session_state.off_slots,
                supplementary_schedules=st.session_state.supplementary_schedules,
                biweekly_choice=st.session_state.get('biweekly_choice')
            )
            st.rerun()

# 탭 인덱스 9: 스케줄 검증
with tabs[9]:
    st.subheader("✅ 스케줄 검증")
    st.caption("정답 목록과 실제 생성된 스케줄을 비교합니다. 공휴일/교수 휴진으로 인한 누락은 자동으로 사유를 표시합니다.")

    if st.session_state.current_df_all.empty:
        st.warning("먼저 사이드바에서 '설정 및 공휴일 확정 적용'을 눌러 스케줄을 생성해주세요.")
    else:
        try:
            verify_result = verify_schedule(
                st.session_state.current_df_all,
                st.session_state.week_count,
                st.session_state.base_date,
                user_holidays,
                st.session_state.off_slots,
                supplementary_schedules=st.session_state.supplementary_schedules,
                assignments=st.session_state.assignments,
                task_date_overrides=st.session_state.get("cpsat_task_date_overrides", {}),
                task_time_overrides=st.session_state.get("cpsat_task_time_overrides", {}),
                shifted_tasks=st.session_state.get("cpsat_shifted_tasks", []),
                original_df_dates=st.session_state.get("cpsat_original_df_dates", {}),
                master_schedules=st.session_state.master_schedules,
                skipped_task_ids=st.session_state.get("cpsat_skipped_obs", []),
                prev_month_task_ids=st.session_state.get("cpsat_skipped_prev_month_obs", []),
                biweekly_choice=st.session_state.get("biweekly_choice"),
            )

            # 전체 요약
            total_days = 0
            clean_days = 0
            total_missing_real = 0  # 사유 없는 진짜 누락
            total_missing_explained = 0  # 사유 있는 누락
            total_extra_real = 0  # 사유 없는 진짜 추가
            total_extra_explained = 0  # 보충진료 등 사유 있는 추가
            total_unassigned = 0  # 미배정 task 수
            total_skipped = 0     # 판정참관 제외룰로 '안 하기로 한' task 수 (미배정 아님)
            total_prev_skipped = 0  # 이전 달 참관 중 배정 안 한 것 (미배정 아님)
            for week_num, days in verify_result.items():
                for day_name, r in days.items():
                    total_days += 1
                    if not r['missing'] and not r['extra'] and not r['unassigned']:
                        clean_days += 1
                    total_skipped += len(r.get('skipped', []))
                    total_prev_skipped += len(r.get('prev_skipped', []))
                    for m in r['missing']:
                        if m['reason']:
                            total_missing_explained += 1
                        else:
                            total_missing_real += 1
                    for e in r['extra']:
                        if e['reason']:
                            total_extra_explained += 1
                        else:
                            total_extra_real += 1
                    total_unassigned += len(r['unassigned'])

            c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
            c1.metric("정상 일수", f"{clean_days}/{total_days}")
            c2.metric("진짜 누락", total_missing_real)
            c3.metric("사유 있는 누락", total_missing_explained)
            c4.metric("추가된 항목", f"{total_extra_real} (+{total_extra_explained})", help="앞: 사유 없는 진짜 추가 / 괄호: 보충진료 등 사유 있는 추가")
            c5.metric("🚨 미배정 task", total_unassigned, help="task는 생성됐으나 자동/수동 배정에서 빠진 task. 절대원칙 위반: '모든 task는 빠짐없이 배정' + '한 세션 1 task'")
            c6.metric("🚫 판정참관 제외", total_skipped, help="판정참관 배정 제외 대상자가 건증 판정을 받아, 룰에 따라 의도적으로 배정하지 않은 참관. 미배정이 아니며 조치가 필요 없습니다.")
            c7.metric("📅 이전달 참관 미배정", total_prev_skipped, help="짝(차리/판정)이 지난달이라 이번 달엔 참관만 남은 task 중, '이전 달 참관 배정 안 함' 옵션으로 배정하지 않은 것. 미배정이 아니며 조치가 필요 없습니다.")

            if total_unassigned > 0:
                st.error(f"🚨 **미배정 task가 {total_unassigned}개 있습니다.** 슬롯 부족으로 배정되지 못한 task입니다. 절대원칙 위반 — 수동 배정이 필요합니다.")
            if total_missing_real == 0 and total_extra_real == 0 and total_unassigned == 0:
                st.success(f"🎉 검증 결과: 모든 항목이 정답 목록과 일치하고 미배정도 없습니다! (사유 있는 누락 {total_missing_explained}건, 사유 있는 추가 {total_extra_explained}건은 정상 동작)")
            elif total_missing_real == 0 and total_unassigned == 0:
                st.info(f"ℹ️ 진짜 누락/미배정은 없지만, 정답 목록에 없는 항목이 {total_extra_real}개 추가되어 있습니다. (이동된 task일 가능성 있음)")
            elif total_missing_real > 0:
                st.error(f"❌ 정답 목록에 있어야 하는데 사유 없이 누락된 항목이 {total_missing_real}개 있습니다.")

            st.markdown("---")

            # === [신규] 미배정 task 별도 섹션 (가장 먼저, 가장 눈에 띄게) ===
            if total_unassigned > 0:
                st.markdown("### 🚨 미배정 task 목록")
                st.caption("아래 task들은 스케줄에는 생성되었지만 배정받은 전공의가 없습니다. 스케줄 배정 탭에서 수동으로 배정해주세요.")
                for week_num in sorted(verify_result.keys()):
                    week_has_unassigned = False
                    for day_name in ["월", "화", "수", "목", "금"]:
                        if verify_result[week_num][day_name]['unassigned']:
                            week_has_unassigned = True
                            break
                    if not week_has_unassigned:
                        continue
                    st.markdown(f"**📅 {week_num}주차**")
                    for day_name in ["월", "화", "수", "목", "금"]:
                        r = verify_result[week_num][day_name]
                        if r['unassigned']:
                            st.markdown(f"<div style='padding:6px 10px; margin:4px 0; background:#FDF2F2; border-left:4px solid #c0392b; border-radius:3px;'><b>{day_name}요일 ({r['date']})</b> — 미배정 {len(r['unassigned'])}개</div>", unsafe_allow_html=True)
                            for t in r['unassigned']:
                                st.markdown(f"<div style='margin-left:40px; color:#c0392b;'>🚨 {t}</div>", unsafe_allow_html=True)
                st.markdown("---")

            # === 판정참관 제외 목록 (룰에 따른 의도적 미배정 — 조치 불필요) ===
            if total_skipped > 0:
                with st.expander(f"🚫 판정참관 제외 {total_skipped}개 — 룰에 따라 배정하지 않음 (조치 불필요)"):
                    st.caption(
                        "'건증 판정참관 배정 제외 대상'으로 지정된 전공의가 그 짝의 건증 판정을 받아, "
                        "판정참관을 의도적으로 배정하지 않은 것입니다. 미배정과 무관하며 수동 배정이 필요 없습니다."
                    )
                    for week_num in sorted(verify_result.keys()):
                        if not any(verify_result[week_num][d].get('skipped') for d in ["월", "화", "수", "목", "금"]):
                            continue
                        st.markdown(f"**📅 {week_num}주차**")
                        for day_name in ["월", "화", "수", "목", "금"]:
                            r = verify_result[week_num][day_name]
                            if r.get('skipped'):
                                st.markdown(
                                    f"<div style='padding:6px 10px; margin:4px 0; background:#F4F4F4; border-left:4px solid #888; border-radius:3px;'>"
                                    f"<b>{day_name}요일 ({r['date']})</b> — 제외 {len(r['skipped'])}개</div>",
                                    unsafe_allow_html=True)
                                for t in r['skipped']:
                                    st.markdown(f"<div style='margin-left:40px; color:#666;'>🚫 {t}</div>", unsafe_allow_html=True)
                st.markdown("---")

            # === 이전 달 참관 미배정 목록 (옵션에 따른 의도적 미배정 — 조치 불필요) ===
            if total_prev_skipped > 0:
                with st.expander(f"📅 이전달 참관 미배정 {total_prev_skipped}개 — 옵션에 따라 배정 안 함 (조치 불필요)"):
                    st.caption(
                        "짝(차리/판정)이 지난달이라 이번 달엔 참관만 남은 task입니다. "
                        "'이전 달 참관 배정 안 해도 됨' 옵션에 따라 배정하지 않은 것으로, 진짜 미배정이 아니며 수동 배정이 필요 없습니다."
                    )
                    for week_num in sorted(verify_result.keys()):
                        if not any(verify_result[week_num][d].get('prev_skipped') for d in ["월", "화", "수", "목", "금"]):
                            continue
                        st.markdown(f"**📅 {week_num}주차**")
                        for day_name in ["월", "화", "수", "목", "금"]:
                            r = verify_result[week_num][day_name]
                            if r.get('prev_skipped'):
                                st.markdown(
                                    f"<div style='padding:6px 10px; margin:4px 0; background:#EEF3F8; border-left:4px solid #4a90d9; border-radius:3px;'>"
                                    f"<b>{day_name}요일 ({r['date']})</b> — 이전달 참관 {len(r['prev_skipped'])}개</div>",
                                    unsafe_allow_html=True)
                                for t in r['prev_skipped']:
                                    st.markdown(f"<div style='margin-left:40px; color:#4a6a8a;'>📅 {t}</div>", unsafe_allow_html=True)
                st.markdown("---")

            # 주차별 상세
            st.markdown("### 📋 주차별 상세 (누락/추가)")
            for week_num in sorted(verify_result.keys()):
                with st.expander(f"📅 {week_num}주차 상세", expanded=(total_missing_real > 0 or total_extra_real > 0)):
                    for day_name in ["월", "화", "수", "목", "금"]:
                        r = verify_result[week_num][day_name]
                        has_real_missing = any(m['reason'] is None for m in r['missing'])
                        has_real_extra = any(e['reason'] is None for e in r['extra'])
                        has_extra = len(r['extra']) > 0

                        if not r['missing'] and not has_extra and not r.get('moved_in') and not r.get('moved_out'):
                            st.markdown(f"<div style='padding:6px 10px; margin:4px 0; background:#E8F8F5; border-left:4px solid #1abc9c; border-radius:3px;'>✅ <b>{day_name}요일 ({r['date']})</b> — 정답 {r['total_expected']}개 / 실제 {r['total_present']}개, 모두 일치</div>", unsafe_allow_html=True)
                        else:
                            border_color = "#e74c3c" if has_real_missing else ("#f39c12" if has_real_extra else "#95a5a6")
                            st.markdown(f"<div style='padding:6px 10px; margin:4px 0; background:#FDF2F2; border-left:4px solid {border_color}; border-radius:3px;'><b>{day_name}요일 ({r['date']})</b> — 정답 {r['total_expected']}개 / 실제 {r['total_present']}개</div>", unsafe_allow_html=True)

                            # 진짜 누락 (사유 없음)
                            real_missing = [m for m in r['missing'] if not m['reason']]
                            if real_missing:
                                st.markdown(f"<div style='margin-left:20px; color:#c0392b;'>❌ <b>진짜 누락 ({len(real_missing)}개) — 확인 필요</b></div>", unsafe_allow_html=True)
                                for m in real_missing:
                                    st.markdown(f"<div style='margin-left:40px; color:#c0392b;'>- {m['name']}</div>", unsafe_allow_html=True)

                            # 사유 있는 누락
                            explained_missing = [m for m in r['missing'] if m['reason']]
                            if explained_missing:
                                st.markdown(f"<div style='margin-left:20px; color:#7f8c8d;'>ℹ️ 사유 있는 누락 ({len(explained_missing)}개)</div>", unsafe_allow_html=True)
                                for m in explained_missing:
                                    st.markdown(f"<div style='margin-left:40px; color:#7f8c8d;'>- {m['name']} → <i>{m['reason']}</i></div>", unsafe_allow_html=True)

                            # 진짜 추가 (사유 없음)
                            real_extra = [e for e in r['extra'] if not e['reason']]
                            if real_extra:
                                st.markdown(f"<div style='margin-left:20px; color:#d68910;'>⚠️ 추가된 항목 ({len(real_extra)}개) — 정답 목록에 없음 (이동/추적 안됨)</div>", unsafe_allow_html=True)
                                for e in real_extra:
                                    st.markdown(f"<div style='margin-left:40px; color:#d68910;'>+ {e['name']}</div>", unsafe_allow_html=True)

                            # 사유 있는 추가 (보충진료 등)
                            explained_extra = [e for e in r['extra'] if e['reason']]
                            if explained_extra:
                                st.markdown(f"<div style='margin-left:20px; color:#7f8c8d;'>ℹ️ 사유 있는 추가 ({len(explained_extra)}개)</div>", unsafe_allow_html=True)
                                for e in explained_extra:
                                    st.markdown(f"<div style='margin-left:40px; color:#7f8c8d;'>+ {e['name']} → <i>{e['reason']}</i></div>", unsafe_allow_html=True)

                            # 🔄 이동 들어옴 (다른 날짜에서 이 날짜로 옴)
                            moved_in = r.get('moved_in', [])
                            if moved_in:
                                st.markdown(f"<div style='margin-left:20px; color:#2980b9;'>🔄 다른 날짜에서 이동 들어옴 ({len(moved_in)}개)</div>", unsafe_allow_html=True)
                                for mi in moved_in:
                                    from_info = f"{mi['from_day']}요일" + (f" ({mi['from_date']})" if mi['from_date'] else "")
                                    st.markdown(f"<div style='margin-left:40px; color:#2980b9;'>← {mi['name']} (원래 {from_info}) — <i>{mi['reason']}</i></div>", unsafe_allow_html=True)

                            # 🔄 이동 나감 (이 날짜에서 다른 날짜로 감)
                            moved_out = r.get('moved_out', [])
                            if moved_out:
                                st.markdown(f"<div style='margin-left:20px; color:#16a085;'>🔄 다른 날짜로 이동 나감 ({len(moved_out)}개)</div>", unsafe_allow_html=True)
                                for mo in moved_out:
                                    to_info = f"{mo['to_day']}요일" + (f" ({mo['to_date']})" if mo['to_date'] else "")
                                    st.markdown(f"<div style='margin-left:40px; color:#16a085;'>→ {mo['name']} (→ {to_info}) — <i>{mo['reason']}</i></div>", unsafe_allow_html=True)
        except Exception as e:
            import traceback
            st.error(f"검증 중 오류 발생: {str(e)}")
            st.code(traceback.format_exc())
